#!/usr/bin/env python3
"""Fuehrt die Rechercheergebnisse aus den JSON-Dateien zu einer Lead-Liste zusammen.

Erzeugt leads-adept.csv und leads-adept.xlsx, sortiert nach Branche.
Verwirft dabei alles, was die Qualitaetsregeln verletzt: Leads ohne persoenliche
E-Mail, ohne Quell-URL, sowie Dubletten (gleiche E-Mail).
"""
import csv
import json
import re
import sys
from pathlib import Path

BRANCHEN_REIHENFOLGE = [
    "Automobil & Zulieferer",
    "Maschinenbau & Fertigung",
    "Logistik & Supply Chain",
    "Chemie & Prozessindustrie",
]

# Rollen-/Funktionspostfaecher, die laut Aufgabenstellung nicht zaehlen.
GENERISCHE_PRAEFIXE = {
    "info", "kontakt", "contact", "vertrieb", "sales", "mail", "email", "office",
    "zentrale", "service", "buero", "post", "empfang", "karriere", "jobs",
    "bewerbung", "hr", "personal", "einkauf", "verkauf", "support", "dispo",
    "disposition", "logistik", "auftrag", "bestellung", "anfrage", "presse",
    "marketing", "webmaster", "datenschutz", "impressum", "newsletter",
    "buchhaltung", "rechnung", "finance", "qm", "technik", "team", "willkommen",
    "hallo", "moin", "firma", "company", "export", "import", "lager",
}

EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")

SPALTEN = [
    "Name", "Position", "E-Mail", "Telefon", "Firmenname", "Branche", "Ort",
    "Beschreibung", "Quelle",
]


def ist_persoenlich(email: str) -> bool:
    """True, wenn die Adresse auf eine Person und nicht auf ein Postfach zeigt."""
    if not EMAIL_RE.match(email):
        return False
    lokal = email.split("@", 1)[0].lower()
    if lokal in GENERISCHE_PRAEFIXE:
        return False
    # Zusammengesetzte Funktionsadressen wie info-nord@ oder vertrieb.de@
    for teil in re.split(r"[._-]", lokal):
        if teil in GENERISCHE_PRAEFIXE:
            return False
    return True


def lade_leads(verzeichnis: Path):
    leads, verworfen, quellen = [], [], []
    for pfad in sorted(verzeichnis.glob("*.json")):
        try:
            daten = json.loads(pfad.read_text(encoding="utf-8"))
        except json.JSONDecodeError as fehler:
            print(f"WARNUNG: {pfad.name} ist kein gueltiges JSON ({fehler})", file=sys.stderr)
            continue
        quellen.append(pfad.name)
        branche = daten.get("branche", "")
        for lead in daten.get("leads", []):
            lead.setdefault("branche", branche)
            lead["_datei"] = pfad.name
            leads.append(lead)
        for eintrag in daten.get("verworfen", []):
            eintrag["branche"] = branche
            verworfen.append(eintrag)
    return leads, verworfen, quellen


def pruefe_und_dedupliziere(leads):
    """Filtert Leads ohne persoenliche E-Mail bzw. ohne Quelle und entfernt Dubletten."""
    behalten, aussortiert = [], []
    gesehen = {}
    for lead in leads:
        email = (lead.get("email") or "").strip()
        quelle = (lead.get("email_quelle") or "").strip()
        firma = lead.get("firma", "?")
        if not email:
            aussortiert.append((firma, "keine E-Mail im Ergebnis"))
            continue
        if not ist_persoenlich(email):
            aussortiert.append((firma, f"Funktionsadresse: {email}"))
            continue
        if not quelle.startswith("http"):
            aussortiert.append((firma, f"keine belastbare Quell-URL fuer {email}"))
            continue
        schluessel = email.lower()
        if schluessel in gesehen:
            aussortiert.append((firma, f"Dublette zu {gesehen[schluessel]}"))
            continue
        gesehen[schluessel] = firma
        behalten.append(lead)
    return behalten, aussortiert


def sortierschluessel(lead):
    branche = lead.get("branche", "")
    rang = BRANCHEN_REIHENFOLGE.index(branche) if branche in BRANCHEN_REIHENFOLGE else len(BRANCHEN_REIHENFOLGE)
    return (rang, lead.get("firma", ""), lead.get("name", ""))


def als_zeile(lead):
    quellen = [lead.get("email_quelle", "")]
    for weitere in lead.get("weitere_quellen", []) or []:
        if weitere and weitere not in quellen:
            quellen.append(weitere)
    telefon = (lead.get("telefon") or "").strip() or "nicht öffentlich"
    return {
        "Name": lead.get("name", ""),
        "Position": lead.get("position", ""),
        "E-Mail": lead.get("email", ""),
        "Telefon": telefon,
        "Firmenname": lead.get("firma", ""),
        "Branche": lead.get("branche", ""),
        "Ort": lead.get("ort", ""),
        "Beschreibung": lead.get("beschreibung", ""),
        "Quelle": " | ".join(q for q in quellen if q),
    }


def schreibe_csv(zeilen, ziel: Path):
    with ziel.open("w", newline="", encoding="utf-8-sig") as datei:
        schreiber = csv.DictWriter(datei, fieldnames=SPALTEN, delimiter=";")
        schreiber.writeheader()
        schreiber.writerows(zeilen)


def schreibe_xlsx(zeilen, verworfen, statistik, ziel: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Hinweis: openpyxl fehlt, XLSX wird uebersprungen.", file=sys.stderr)
        return False

    mappe = Workbook()
    blatt = mappe.active
    blatt.title = "Leads"
    blatt.append(SPALTEN)

    kopf_fuellung = PatternFill("solid", start_color="1F3864")
    for zelle in blatt[1]:
        zelle.font = Font(bold=True, color="FFFFFF")
        zelle.fill = kopf_fuellung
        zelle.alignment = Alignment(vertical="center")

    for zeile in zeilen:
        blatt.append([zeile[spalte] for spalte in SPALTEN])

    breiten = {"Name": 24, "Position": 30, "E-Mail": 34, "Telefon": 22,
               "Firmenname": 34, "Branche": 24, "Ort": 24, "Beschreibung": 70,
               "Quelle": 60}
    for index, spalte in enumerate(SPALTEN, start=1):
        blatt.column_dimensions[get_column_letter(index)].width = breiten[spalte]
    for reihe in blatt.iter_rows(min_row=2):
        for zelle in reihe:
            zelle.alignment = Alignment(vertical="top", wrap_text=True)
    blatt.freeze_panes = "A2"
    blatt.auto_filter.ref = blatt.dimensions

    zusammenfassung = mappe.create_sheet("Zusammenfassung")
    zusammenfassung.append(["Branche", "Gefundene Leads", "Verworfene Kandidaten"])
    for zelle in zusammenfassung[1]:
        zelle.font = Font(bold=True, color="FFFFFF")
        zelle.fill = kopf_fuellung
    for branche in BRANCHEN_REIHENFOLGE:
        zusammenfassung.append([branche, statistik["leads"].get(branche, 0),
                                statistik["verworfen"].get(branche, 0)])
    zusammenfassung.append(["Gesamt", sum(statistik["leads"].values()),
                            sum(statistik["verworfen"].values())])
    for spalte, breite in zip("ABC", (34, 18, 24)):
        zusammenfassung.column_dimensions[spalte].width = breite

    verworfen_blatt = mappe.create_sheet("Verworfen")
    verworfen_blatt.append(["Firma", "Ort", "Branche", "Grund"])
    for zelle in verworfen_blatt[1]:
        zelle.font = Font(bold=True, color="FFFFFF")
        zelle.fill = kopf_fuellung
    for eintrag in verworfen:
        verworfen_blatt.append([eintrag.get("firma", ""), eintrag.get("ort", ""),
                                eintrag.get("branche", ""), eintrag.get("grund", "")])
    for spalte, breite in zip("ABCD", (38, 24, 26, 70)):
        verworfen_blatt.column_dimensions[spalte].width = breite

    mappe.save(ziel)
    return True


def main():
    basis = Path(__file__).resolve().parent
    quellverzeichnis = Path(sys.argv[1]) if len(sys.argv) > 1 else basis / "rohdaten"

    leads, verworfen, dateien = lade_leads(quellverzeichnis)
    behalten, aussortiert = pruefe_und_dedupliziere(leads)
    behalten.sort(key=sortierschluessel)
    zeilen = [als_zeile(lead) for lead in behalten]

    statistik = {"leads": {}, "verworfen": {}}
    for lead in behalten:
        branche = lead.get("branche", "")
        statistik["leads"][branche] = statistik["leads"].get(branche, 0) + 1
    for eintrag in verworfen:
        branche = eintrag.get("branche", "")
        statistik["verworfen"][branche] = statistik["verworfen"].get(branche, 0) + 1
    for _, grund in aussortiert:
        pass

    schreibe_csv(zeilen, basis / "leads-adept.csv")
    xlsx_ok = schreibe_xlsx(zeilen, verworfen, statistik, basis / "leads-adept.xlsx")

    print(f"Quelldateien: {len(dateien)}")
    print(f"Leads gesamt: {len(zeilen)}")
    for branche in BRANCHEN_REIHENFOLGE:
        print(f"  {branche}: {statistik['leads'].get(branche, 0)} Leads, "
              f"{statistik['verworfen'].get(branche, 0)} verworfen")
    if aussortiert:
        print(f"Bei der Qualitaetspruefung zusaetzlich entfernt: {len(aussortiert)}")
        for firma, grund in aussortiert:
            print(f"  - {firma}: {grund}")
    print(f"Geschrieben: leads-adept.csv{' und leads-adept.xlsx' if xlsx_ok else ''}")


if __name__ == "__main__":
    main()
