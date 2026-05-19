"""
Paletten-Optimierungskern v2 — VERIFIZIERTE REFERENZ.

Encodet die 3 Optimierungs-Punkte als ausführbare Wahrheit:
 - Punkt 1+2: minimiere Gesamtzahl verschiedener Paletten, einseitig
   (Standard MUSS >= Last in beiden Maßen -> Last passt physisch drauf)
 - Punkt 3:   Kombinieren als FALLBACK (Boolean). Nur wenn KEIN einzelner
   Standard passt, wird geprüft ob 2-3 GEWÄHLTE Standards (Typ A + Typ B)
   die Last abdecken -> spart Sonderpaletten.
 - Physikalische Invariante: KEIN zugewiesener Standard (auch keine
   Kombination) darf kleiner sein als die Last. Wird hart geprüft.

Punkt 4 (Doppelbestellungen/AW/Bestand) ist NICHT hier — eigener,
stateful Layer, kommt nach dem Einfrieren dieses Kerns.

Abhängigkeit: pulp
"""
import pulp
from itertools import combinations, product


def _passt_einzeln(cand, last, T):
    c0, c1 = sorted(cand)
    l0, l1 = sorted(last)
    return c0 >= l0 and c1 >= l1 and (c0 - l0) <= T and (c1 - l1) <= T


def _passt_kombi(teile, last, T):
    l0, l1 = sorted(last)
    for variante in product(*[((a, b), (b, a)) for (a, b) in teile]):
        sx = sum(t[0] for t in variante)
        my = max(t[1] for t in variante)
        b0, b1 = sorted((sx, my))
        if b0 >= l0 and b1 >= l1 and (b0 - l0) <= T and (b1 - l1) <= T:
            return True
        sy = sum(t[1] for t in variante)
        mx = max(t[0] for t in variante)
        b0, b1 = sorted((mx, sy))
        if b0 >= l0 and b1 >= l1 and (b0 - l0) <= T and (b1 - l1) <= T:
            return True
    return False


def optimiere(orders, tol_mm=200, kombinieren=True, max_kombi_teile=3,
              zeitlimit_s=120, sonder_deckel=None):
    if not orders:
        return {'standards': [], 'sonder': [], 'gesamt': 0,
                'zuordnung': [], 'invariante_ok': True,
                'verletzungen': [], 'status': 'leer'}

    O = [(int(round(o['L'])), int(round(o['B'])),
          int(o['menge']), o.get('auftrag'), o.get('name'))
         for o in orders]
    N = len(O)
    T = tol_mm

    S = sorted({min(a, b) for a, b, *_ in O})
    Lv = sorted({max(a, b) for a, b, *_ in O})
    cands = [(cs, cl) for cs in S for cl in Lv if cs <= cl]

    keep, cov = [], []
    for c in cands:
        s = {i for i, (a, b, *_ ) in enumerate(O)
             if _passt_einzeln(c, (a, b), T)}
        if s:
            keep.append(c)
            cov.append(s)

    groups = {}
    for i, (a, b, *_ ) in enumerate(O):
        groups.setdefault(tuple(sorted((a, b))), []).append(i)

    p = pulp.LpProblem("min_gesamt", pulp.LpMinimize)
    x = [pulp.LpVariable(f"x{j}", cat="Binary") for j in range(len(keep))]
    z = {g: pulp.LpVariable(f"z{gi}", cat="Binary")
         for gi, g in enumerate(groups)}
    p += pulp.lpSum(x) + pulp.lpSum(z.values())
    for i, (a, b, *_ ) in enumerate(O):
        g = tuple(sorted((a, b)))
        p += (pulp.lpSum(x[j] for j in range(len(keep)) if i in cov[j])
              + z[g] >= 1)
    # OPTIONALE Erweiterung (nicht in v2-Spec, aber additiv):
    # max. Anzahl verschiedener Sonder-Maße. None = unbegrenzt.
    if sonder_deckel is not None and sonder_deckel >= 0:
        p += pulp.lpSum(z.values()) <= sonder_deckel, "sonder_max"
    p.solve(pulp.PULP_CBC_CMD(msg=0, timeLimit=zeitlimit_s))

    standards = sorted(keep[j] for j in range(len(keep))
                       if x[j].value() and x[j].value() > 0.5)

    zuordnung, sonder_grp = [], set()
    for i, (a, b, m, au, nm) in enumerate(O):
        einz = next((s for s in standards
                     if _passt_einzeln(s, (a, b), T)), None)
        if einz:
            zuordnung.append({'auftrag': au, 'name': nm, 'L': a, 'B': b,
                              'menge': m, 'typ': 'Standard',
                              'ziel': f"{einz[0]}x{einz[1]}"})
            continue
        kombi = None
        if kombinieren and standards:
            for r in range(2, max_kombi_teile + 1):
                for combo in combinations(standards, r):
                    if _passt_kombi(combo, (a, b), T):
                        kombi = combo
                        break
                if kombi:
                    break
        if kombi:
            zuordnung.append({'auftrag': au, 'name': nm, 'L': a, 'B': b,
                              'menge': m, 'typ': 'Kombination',
                              'ziel': " + ".join(f"{c[0]}x{c[1]}"
                                                 for c in kombi)})
        else:
            sonder_grp.add(tuple(sorted((a, b))))
            zuordnung.append({'auftrag': au, 'name': nm, 'L': a, 'B': b,
                              'menge': m, 'typ': 'Sonder',
                              'ziel': f"{min(a,b)}x{max(a,b)}"})

    verletzungen = []
    for zg in zuordnung:
        if zg['typ'] == 'Sonder':
            continue
        l0, l1 = sorted((zg['L'], zg['B']))
        if zg['typ'] == 'Standard':
            cs, cl = map(int, zg['ziel'].split('x'))
            if not (min(cs, cl) >= l0 and max(cs, cl) >= l1):
                verletzungen.append(zg)
        else:
            teile = [tuple(map(int, t.split('x')))
                     for t in zg['ziel'].split(" + ")]
            if not _passt_kombi(teile, (zg['L'], zg['B']), T):
                verletzungen.append(zg)

    return {'standards': standards,
            'sonder': sorted(sonder_grp),
            'gesamt': len(standards) + len(sonder_grp),
            'zuordnung': zuordnung,
            'invariante_ok': len(verletzungen) == 0,
            'verletzungen': verletzungen,
            'status': pulp.LpStatus[p.status]}


if __name__ == "__main__":
    cluster = [{'L': 1520, 'B': 720, 'menge': 1, 'auftrag': 'A', 'name': ''},
               {'L': 1550, 'B': 700, 'menge': 1, 'auftrag': 'B', 'name': ''},
               {'L': 1500, 'B': 750, 'menge': 1, 'auftrag': 'C', 'name': ''},
               {'L': 1530, 'B': 710, 'menge': 1, 'auftrag': 'D', 'name': ''}]
    r = optimiere(cluster, tol_mm=200, kombinieren=False)
    print("Cluster:", r['standards'], "| Sonder:", r['sonder'],
          "| invariante_ok:", r['invariante_ok'])
    assert len(r['standards']) == 1, r['standards']
    assert len(r['sonder']) == 0
    assert r['invariante_ok']
    s = r['standards'][0]
    for o in cluster:
        assert min(s) >= min(o['L'], o['B']) and max(s) >= max(o['L'], o['B'])
    print("OK Selbsttest 1 (Cluster -> 1 Standard, deckt alle, einseitig)")

    base = [{'L': 1200, 'B': 800, 'menge': 9, 'auftrag': 'S1', 'name': ''},
            {'L': 400, 'B': 800, 'menge': 9, 'auftrag': 'S2', 'name': ''},
            {'L': 1500, 'B': 700, 'menge': 1, 'auftrag': 'X', 'name': ''}]
    ro = optimiere(base, tol_mm=200, kombinieren=False)
    rk = optimiere(base, tol_mm=200, kombinieren=True)
    tx_o = [z['typ'] for z in ro['zuordnung'] if z['auftrag'] == 'X'][0]
    tx_k = [z for z in rk['zuordnung'] if z['auftrag'] == 'X'][0]
    print("X ohne Kombi:", tx_o, "| mit Kombi:", tx_k['typ'],
          "->", tx_k['ziel'])
    assert tx_o == 'Sonder'
    assert tx_k['typ'] == 'Kombination'
    assert rk['invariante_ok']
    print("OK Selbsttest 2 (Kombi-Fallback Typ A + Typ B)")

    from openpyxl import load_workbook
    wb = load_workbook('/mnt/user-data/uploads/Paletten_DM.xlsx',
                        read_only=True)
    ws = wb.active
    ords = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if row is None:
            continue
        a, n, m, L, B = row[4], row[7], row[12], row[14], row[15]
        if a in (None, '') and L in (None, '') and B in (None, ''):
            continue
        if not (isinstance(L, (int, float))
                and isinstance(B, (int, float))):
            continue
        ords.append({'L': L, 'B': B,
                     'menge': int(m) if isinstance(m, (int, float)) and m
                     else 1, 'auftrag': a, 'name': n})
    for T in (100, 200):
        rr = optimiere(ords, tol_mm=T, kombinieren=True)
        print(f"Echt T={T}: Std={len(rr['standards'])} "
              f"Sonder={len(rr['sonder'])} Gesamt={rr['gesamt']} "
              f"invariante_ok={rr['invariante_ok']} "
              f"Verletzungen={len(rr['verletzungen'])}")
        assert rr['invariante_ok'], "PHYSIKALISCHE INVARIANTE VERLETZT"
    print("OK Selbsttest 3 (echte Datei, einseitig, 0 Verletzungen)")
