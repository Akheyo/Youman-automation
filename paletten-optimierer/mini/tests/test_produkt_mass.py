"""Pflicht-Tests fuer den Domain-Fix: Excel P-Maße = Paletten-Maße,
Produkt-Maß = Paletten-Maß minus PALETTEN_AUFSCHLAG_MM (50 mm).

Vier harte Asserts. Build rot wenn einer faellt.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

HIER = Path(__file__).resolve().parent
sys.path.insert(0, str(HIER.parent))

import openpyxl  # noqa: E402

from import_excel import importiere, PALETTEN_AUFSCHLAG_MM  # noqa: E402
from optimierer_kern import optimiere  # noqa: E402


# Spec-Soll fuer die echte Datei
ECHTE_DATEI = HIER / "fixtures" / "Paletten_DM.xlsx"
SOLL_ECHT = {
    "mit_mass": 469,
    "ohne_mass": 39,
    "paletten_gesamt": 2724,
    "normalisierte_masse": 158,
}


def _baue_minimal_xlsx(tmp: Path, zeilen: list[dict]) -> Path:
    """Baut eine minimale Mini-Excel mit Header in Z.5 (wie die echte
    Struktur). Jede Datenzeile in zeilen ist ein dict mit den Keys
    'auftrag','name','artikelnr','menge','L','B'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Testfirma")
    ws.cell(row=3, column=1, value="verplante Maschinen")
    # Header in Z.5 mit Spalten-Anordnung wie echte Datei
    ws.cell(row=5, column=1,  value="Maschine")
    ws.cell(row=5, column=5,  value="Auftrag")
    ws.cell(row=5, column=8,  value="Name")
    ws.cell(row=5, column=10, value="Artikelnummer")
    ws.cell(row=5, column=13, value="Menge")
    ws.cell(row=5, column=15, value="P-Länge")
    ws.cell(row=5, column=16, value="P-Breite")
    # Daten ab Z.7, jede zweite Zeile (Leerzeile dazwischen)
    r = 7
    for z in zeilen:
        ws.cell(row=r, column=5,  value=z.get("auftrag", ""))
        ws.cell(row=r, column=8,  value=z.get("name", ""))
        ws.cell(row=r, column=10, value=z.get("artikelnr", ""))
        ws.cell(row=r, column=13, value=z.get("menge", 1))
        if z.get("L") is not None:
            ws.cell(row=r, column=15, value=z["L"])
        if z.get("B") is not None:
            ws.cell(row=r, column=16, value=z["B"])
        r += 2
    wb.save(tmp)
    return tmp


def _tmp_xlsx(name: str) -> Path:
    """Plattform-neutraler Pfad fuer eine temporaere xlsx."""
    import tempfile
    return Path(tempfile.gettempdir()) / f"paletten_test_{name}.xlsx"


def _cleanup(p: Path) -> None:
    """Robust loeschen — auf Windows kann Datei kurz noch belegt sein."""
    try:
        if p.exists():
            p.unlink()
    except (PermissionError, OSError):
        pass  # Test-Erfolg darf nicht am Cleanup haengen


# ---------------------------------------------------------------------
# a) P-L=1450, P-B=850 -> L=1400, B=800 beim Optimierer
# ---------------------------------------------------------------------
def test_a_subtraktion_50_mm():
    assert PALETTEN_AUFSCHLAG_MM == 50, "Konstante muss 50 sein (Draht-Mueller)"
    tmp = _tmp_xlsx("a_subtraktion")
    try:
        _baue_minimal_xlsx(tmp, [
            {"auftrag": "A1", "name": "K1", "artikelnr": "ART1",
             "menge": 1, "L": 1450, "B": 850},
        ])
        dat = importiere(tmp)
        assert len(dat["mit_mass"]) == 1
        e = dat["mit_mass"][0]
        assert e["laenge"] == 1400, (
            f"Produkt-L = {e['laenge']}, erwartet 1400 (1450 − 50)"
        )
        assert e["breite"] == 800, (
            f"Produkt-B = {e['breite']}, erwartet 800 (850 − 50)"
        )
        assert e["palette_L_excel"] == 1450, "Roh-P-L muss erhalten bleiben"
        assert e["palette_B_excel"] == 850, "Roh-P-B muss erhalten bleiben"

        orders = [{"L": e["laenge"], "B": e["breite"], "menge": e["anzahl"],
                   "auftrag": e["auftrag"], "name": e["name"]}]
        r = optimiere(orders, tol_mm=200, kombinieren=True)
        assert r["zuordnung"][0]["typ"] == "Standard", (
            f"Erwartet Typ=Standard, bekam {r['zuordnung'][0]['typ']!r}"
        )
        assert r["zuordnung"][0]["L"] == 1400
        assert r["zuordnung"][0]["B"] == 800
        assert r["invariante_ok"], "Invariante muss OK sein"
        print(f"  OK — P-L=1450/P-B=850 -> Produkt 1400×800, "
              f"Roh-Werte erhalten, Standard, Invariante OK.")
    finally:
        _cleanup(tmp)


# ---------------------------------------------------------------------
# b) P-L=40, P-B=850 -> produkt_L=-10 -> ohne_mass-Bucket, kein Crash
# ---------------------------------------------------------------------
def test_b_zu_kleiner_palettenwert_landet_in_bucket():
    tmp = _tmp_xlsx("b_kleiner")
    try:
        _baue_minimal_xlsx(tmp, [
            {"auftrag": "OHNE", "name": "K0", "artikelnr": "ART0",
             "menge": 1, "L": 40, "B": 850},
            {"auftrag": "MIT",  "name": "K1", "artikelnr": "ART1",
             "menge": 2, "L": 1200, "B": 800},
        ])
        dat = importiere(tmp)
        auftrags_mit = {p["auftrag"] for p in dat["mit_mass"]}
        auftrags_ohne = {p["auftrag"] for p in dat["ohne_mass"]}
        print(f"  mit_mass={auftrags_mit}, ohne_mass={auftrags_ohne}")
        assert "OHNE" in auftrags_ohne, (
            "Auftrag mit P-L=40 (Produkt-L=-10) MUSS in ohne_mass landen"
        )
        assert "OHNE" not in auftrags_mit
        assert "MIT" in auftrags_mit

        orders = [{"L": p["laenge"], "B": p["breite"], "menge": p["anzahl"],
                   "auftrag": p["auftrag"], "name": p["name"]}
                  for p in dat["mit_mass"]]
        r = optimiere(orders, tol_mm=200, kombinieren=True)
        assert r["invariante_ok"]
        print(f"  OK — Auftrag mit P-L=40 landet in 'Maß fehlt', kein Crash.")
    finally:
        _cleanup(tmp)


# ---------------------------------------------------------------------
# c) Echte Datei: Soll-Werte 469/39/2724/158 + tol=200 -> 38/0
# ---------------------------------------------------------------------
def test_c_echte_datei_soll_werte():
    if not ECHTE_DATEI.exists():
        print(f"  Skipped — echte Datei nicht in {ECHTE_DATEI.relative_to(HIER.parent)}")
        return
    dat = importiere(ECHTE_DATEI)
    assert len(dat["mit_mass"]) == SOLL_ECHT["mit_mass"], (
        f"mit_mass = {len(dat['mit_mass'])} != Soll {SOLL_ECHT['mit_mass']}"
    )
    assert len(dat["ohne_mass"]) == SOLL_ECHT["ohne_mass"]
    assert dat["paletten_gesamt"] == SOLL_ECHT["paletten_gesamt"]
    assert dat["unique_normalisierte_masse"] == SOLL_ECHT["normalisierte_masse"]
    # Bei tol=200 sollten 38 Standards, 0 Sonder, Invariante OK rauskommen
    orders = [{"L": p["laenge"], "B": p["breite"], "menge": p["anzahl"],
               "auftrag": p["auftrag"], "name": p["name"]}
              for p in dat["mit_mass"]]
    r = optimiere(orders, tol_mm=200, kombinieren=True)
    print(f"  echte Datei tol=200: std={len(r['standards'])}, "
          f"sonder={len(r['sonder'])}, invariante_ok={r['invariante_ok']}")
    assert r["invariante_ok"], "Invariante MUSS OK sein"
    assert len(r["standards"]) == 38, (
        f"Standards = {len(r['standards'])}, erwartet 38"
    )
    assert len(r["sonder"]) == 0, (
        f"Sonder = {len(r['sonder'])}, erwartet 0"
    )


# ---------------------------------------------------------------------
# d) Snapshot: Kombi-Anzeige enthaelt "🔗 Kombination" + "+"
# ---------------------------------------------------------------------
def test_d_anzeige_kombination_format():
    """Konstruiere ein Mini-Szenario, in dem Kombi getriggert wird,
    und pruefe dass die UI-Render-Funktion render_zuord_table() den
    String '🔗 Kombination' und die zwei Standards mit '+' liefert.

    Wichtig: wir importieren _render (reine Render-Funktion ohne
    Streamlit-Side-Effects) statt app_mini — damit der Test auch
    auf einem CI-Runner ohne Streamlit-Runtime sauber durchlaeuft."""
    from _render import render_zuord_table  # noqa: E402

    # Drei Aufträge: S1 (1200×800), S2 (400×800), X (1500×700)
    orders = [
        {"L": 1200, "B": 800, "menge": 9, "auftrag": "S1", "name": ""},
        {"L":  400, "B": 800, "menge": 9, "auftrag": "S2", "name": ""},
        {"L": 1500, "B": 700, "menge": 1, "auftrag": "X",  "name": ""},
    ]
    r = optimiere(orders, tol_mm=200, kombinieren=True)
    kombi = [z for z in r["zuordnung"] if z["typ"] == "Kombination"]
    assert kombi, "Setup-Erwartung: mit Kombi muss X als Kombination kommen"
    assert " + " in kombi[0]["ziel"]

    # Render erwartet artikelnummer + palette_L/B_excel pro Eintrag
    for z in r["zuordnung"]:
        z.setdefault("artikelnummer", "")
        z.setdefault("palette_L_excel", None)
        z.setdefault("palette_B_excel", None)

    html = render_zuord_table(r)
    assert "🔗 Kombination" in html, "Render muss '🔗 Kombination' enthalten"
    assert "+" in html and " mm" in html, "Maße mit '+' verbunden"
    assert 'background:#eff6ff' in html, (
        "Kombi-Zeilen muessen optisch hervorgehoben sein (hellblau)"
    )
    print(f"  OK — Kombi wird als '🔗 Kombination' angezeigt, "
          f"Maße mit '+', Zeile hellblau hinterlegt.")


# ---------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------
if __name__ == "__main__":
    tests = [
        test_a_subtraktion_50_mm,
        test_b_zu_kleiner_palettenwert_landet_in_bucket,
        test_c_echte_datei_soll_werte,
        test_d_anzeige_kombination_format,
    ]
    erfolge = 0
    for fn in tests:
        try:
            fn()
            print(f"OK  {fn.__name__}")
            erfolge += 1
        except AssertionError as e:
            print(f"FAIL {fn.__name__}: {e}")
        except Exception as e:
            print(f"ERR  {fn.__name__}: {type(e).__name__}: {e}")
            import traceback; traceback.print_exc()
    print()
    print("=" * 60)
    print(f"{erfolge}/{len(tests)} Produkt-Maß-Tests bestanden.")
    print("=" * 60)
    sys.exit(0 if erfolge == len(tests) else 1)
