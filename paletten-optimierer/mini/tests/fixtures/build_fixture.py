"""Erzeugt die Pflicht-Test-Fixture für die Mini-App.

Die Fixture wird durch Code aufgebaut — KEINE externe/echte Datei
im Repo. Struktur ist die SCHWIERIGE Variante, die bisher alles
gebrochen hat:

- Zeile 1: Firmenname
- Zeile 2: leer
- Zeile 3: Metazeile "verplante Maschinen" + Datum in Spalte K
- Zeile 4: leer
- Zeile 5: HEADER (Auftrag, Name, Artikelnummer, Menge, Stck Pal,
           P-Länge, P-Breite, P-Höhe, P-Anzahl, ...)
- Zeile 6: leer
- Zeile 7, 9, 11, ...: Datenzeilen mit LEERZEILEN dazwischen

Spalte 'Menge' = echte Palettenanzahl pro Auftrag.
Spalte 'P-Anzahl' ist ABLENKUNG (meist 1) — darf nicht gemappt werden.

Daten enthalten alle Stolperfallen:
- Auftrag mit Drehung 800×1200 + 1200×800 (Orientierung)
- 1 Auftrag mit LEEREN L/B  → Bucket "Maß fehlt"
- 1 Auftragsnummer DOPPELT mit verschiedenen Positionen (/001, /002)
- Cluster mit von Hand nachvollziehbarem Optimum

ERWARTETES OPTIMUM (zum Test verankern):
- 9 Aufträge mit Maß
- 1 Auftrag ohne Maß
- Σ Menge (mit Maß) = 41
- 8 unique kanonische Maße (max, min) — Drehung verschmilzt
- Bei tol=100, kombi=3, zweiseitig: 5 Maße gesamt
  (4 echte Cluster + 1 isolierter Auftrag, kann Standard oder Sonder
   sein; ILP-Optimum = 5)
"""
from __future__ import annotations

from pathlib import Path

import openpyxl


# Spalten-Anordnung 1:1 wie die echte Draht-Müller-Excel:
# A=Maschine B=KW Fertigung C=Nr. D=KW Lief E=Auftrag F=ANr G=AK
# H=Name (I leer) J=Artikelnummer (K leer) L=Masche/Bemerkung
# M=Menge N=Stck Pal O=P-Länge P=P-Breite Q=P-Höhe R=P-Anzahl
HEADER_ROW_VALUES = [
    "Maschine",        # A
    "KW Fertigung",    # B
    "Nr.",             # C
    "KW Lief",         # D
    "Auftrag",         # E
    "ANr",             # F
    "AK",              # G
    "Name",            # H
    None,              # I leer
    "Artikelnummer",   # J
    None,              # K leer
    "Masche/Bemerkung",  # L
    "Menge",           # M  ← Palettenanzahl (richtig)
    "Stck Pal",        # N  ← Stückzahl pro Palette
    "P-Länge",         # O
    "P-Breite",        # P
    "P-Höhe",          # Q
    "P-Anzahl",        # R  ← ABLENKUNG, darf nicht gemappt werden
]

# (auftrag, kunde, artikelnummer, menge, stck_pal, laenge, breite, hoehe, p_anzahl, kommentar)
# leeres laenge/breite = Maß-fehlt-Fall
DATEN = [
    # Cluster A: 1200x800 ± 100, eine Drehung (800x1200) enthalten
    ("110001", "VDL Jansen bv",      "ART-A1-1200x800", 10, 350, 1200,  800,  900, 1, "ClusterA"),
    ("110002", "Müller GmbH",        "ART-A2-1150x750",  5, 100, 1150,  750,  900, 1, "ClusterA"),
    ("110003", "Bressers Metaal BV", "ART-A3-DRH",       3, 125,  800, 1200,  900, 1, "ClusterA-Drehung"),
    # Cluster B: 1500x1000 ± 50
    ("110004", "Big Dutchman",       "ART-B1-1500x1000", 8, 112, 1500, 1000, 1200, 1, "ClusterB"),
    ("110005", "GI-RO Technik GmbH", "ART-B2-1450x950",  6,  38, 1450,  950, 1200, 1, "ClusterB"),
    # Cluster C: isoliert 2000x600
    ("110006", "Hellmann Poultry",   "ART-C1-2000x600",  4, 100, 2000,  600, 1500, 1, "ClusterC"),
    # Sonder-Kandidat: weit weg von allem
    ("110007", "R. Inauen AG",       "ART-X1-3500x2200", 1,  50, 3500, 2200, 1800, 1, "Isoliert"),
    # Doppelte AW-Nummer mit verschiedenen Positionen
    ("110008", "Urban GmbH /001",    "ART-D1-1080x680",  2,  50, 1080,  680,  600, 1, "AW-doppelt-Pos001"),
    ("110008", "Urban GmbH /002",    "ART-D2-1100x700",  2,  50, 1100,  700,  600, 1, "AW-doppelt-Pos002"),
    # Auftrag ohne L/B-Werte → Bucket "Maß fehlt"
    ("110009", "Bosch Rexroth",      "ART-E1-NOMASS",    1, 100, None, None, None, 1, "Maß-fehlt"),
]


def baue_fixture(ziel: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabelle1"

    # Z1: Firmenname (typischer Excel-Export-Stolperfall)
    ws.cell(row=1, column=1, value="Test Müller GmbH & Co. KG")
    # Z2: leer
    # Z3: Metazeile
    ws.cell(row=3, column=1, value="verplante Maschinen")
    ws.cell(row=3, column=11, value="Datum '07.05.26' 20:53, Seite 1")
    # Z4: leer

    # Z5: Header
    for col_idx, val in enumerate(HEADER_ROW_VALUES, start=1):
        ws.cell(row=5, column=col_idx, value=val)
    # Z6: leer

    # Datenzeilen ab Z7, jede zweite Zeile (Z8/Z10/Z12 ... bleiben leer)
    aktuelle_zeile = 7
    for (auftrag, name, artikel, menge, stck, L, B, H, panz, _kommentar) in DATEN:
        ws.cell(row=aktuelle_zeile, column=1,  value="Paletten")   # A: Maschine
        ws.cell(row=aktuelle_zeile, column=2,  value=2619)         # B: KW Fertigung
        ws.cell(row=aktuelle_zeile, column=4,  value=2620)         # D: KW Lief
        ws.cell(row=aktuelle_zeile, column=5,  value=auftrag)      # E: Auftrag
        ws.cell(row=aktuelle_zeile, column=6,  value=0)            # F: ANr
        ws.cell(row=aktuelle_zeile, column=8,  value=name)         # H: Name
        ws.cell(row=aktuelle_zeile, column=10, value=artikel)      # J: Artikelnummer
        ws.cell(row=aktuelle_zeile, column=13, value=menge)        # M: Menge
        ws.cell(row=aktuelle_zeile, column=14, value=stck)         # N: Stck Pal
        if L is not None:
            ws.cell(row=aktuelle_zeile, column=15, value=L)         # O: P-Länge
        if B is not None:
            ws.cell(row=aktuelle_zeile, column=16, value=B)         # P: P-Breite
        if H is not None:
            ws.cell(row=aktuelle_zeile, column=17, value=H)         # Q: P-Höhe
        ws.cell(row=aktuelle_zeile, column=18, value=panz)         # R: P-Anzahl (Ablenkung)
        aktuelle_zeile += 2  # Leerzeile dazwischen

    ziel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ziel)


# ============================================================
# Erwartete Werte für die Pflicht-Asserts
# ============================================================
ERWARTET_HEADER_ZEILE = 5
ERWARTET_MIT_MASS = sum(1 for d in DATEN if d[5] is not None and d[6] is not None)   # 9
ERWARTET_OHNE_MASS = sum(1 for d in DATEN if d[5] is None or d[6] is None)            # 1
ERWARTET_PALETTEN_SUMME = sum(d[3] for d in DATEN if d[5] is not None and d[6] is not None)  # 41
# Kanonisierung: (max(L,B), min(L,B)) — Drehung verschmilzt
_canon = {(max(d[5], d[6]), min(d[5], d[6])) for d in DATEN if d[5] is not None and d[6] is not None}
ERWARTET_UNIQUE_KANON = len(_canon)   # 8

# Kandidaten = volles Gitter (S_vals × L_vals mit cs<=cl)
_s = sorted({min(d[5], d[6]) for d in DATEN if d[5] is not None and d[6] is not None})
_l = sorted({max(d[5], d[6]) for d in DATEN if d[5] is not None and d[6] is not None})
ERWARTET_KANDIDATEN = sum(1 for cs in _s for cl in _l if cs <= cl)   # 57

# Optimum (zu Fuss bei tol=100, kombi=3, zweiseitig)
# Coverage zweiseitig: Standard (Ls,Ws) deckt Auftrag (La,Wa) wenn
#   |Ls-La|<=tol UND |Ws-Wa|<=tol. Standard darf KLEINER sein als
#   Auftrag (Last hängt theoretisch über; das ist genau die User-
#   definierte zweiseitige Coverage).
#
# Damit:
# - Standard (1100,700) deckt mit tol=100 alle Maße mit
#   L in [1000,1200] UND B in [600,800]:
#     (1200,800) ✓, (1150,750) ✓, (1080,680) ✓, (1100,700) ✓
#     plus die Drehung 800×1200 (= kanonisch 1200,800) ✓
#   → deckt 5 Aufträge (110001, 110002, 110003, 110008/001, 110008/002)
# - Standard (1500,1000) deckt (1500,1000) und (1450,950) → 2 Aufträge
# - (2000,600) und (3500,2200) sind isoliert (keine Coverage durch
#   andere) → entweder als Standard oder Sonder (gleicher Score)
# Optimum: 2 echte Cluster + 2 isolierte = 4 Maße gesamt.
ERWARTET_GESAMT = 4


if __name__ == "__main__":
    ziel = Path(__file__).resolve().parent / "pflicht_fixture.xlsx"
    baue_fixture(ziel)
    print(f"Fixture erzeugt: {ziel} ({ziel.stat().st_size} bytes)")
    print(f"Erwartete Werte (zum Anker im Test):")
    print(f"  Header in Datei-Zeile:    {ERWARTET_HEADER_ZEILE}")
    print(f"  mit_mass:                 {ERWARTET_MIT_MASS}")
    print(f"  ohne_mass:                {ERWARTET_OHNE_MASS}")
    print(f"  Σ Menge (mit Maß):       {ERWARTET_PALETTEN_SUMME}")
    print(f"  Unique kanon. Maße:       {ERWARTET_UNIQUE_KANON}")
    print(f"  Kandidaten (volles Gitter): {ERWARTET_KANDIDATEN}")
    print(f"  Optimum (gesamt):         {ERWARTET_GESAMT}")
