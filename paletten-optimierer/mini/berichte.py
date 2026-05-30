"""Berichte-Modul (Spec Berichte-Tab §1-§7).

Erzeugt 5 Berichtstypen + verwaltet das lokale Archiv:
  1. PDF-Optimierungs-Report (jsPDF-Aequivalent: fpdf2)
  2. Excel-Zuordnung (SheetJS-Aequivalent: openpyxl)
  3. Bestandsbericht (PDF oder Excel, mit Filtern)
  4. Wirtschaftlichkeitsbericht (PDF mit Trend-Chart aus PIL)
  5. Lieferanten-Bestelldokument (PDF + Eintrag in beschaffungen)

Persistenz:
  Berichts-Dateien:   %APPDATA%/PalettenMini/berichte/
  Archiv-JSON:        %APPDATA%/PalettenMini/berichte_archiv.json
  ENV-Override:       PALETTENMINI_BERICHTE_DIR
"""
from __future__ import annotations

import io
import json
import os
import sys
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from fpdf import FPDF


# ---------------------------------------------------------------------------
# Pfade
# ---------------------------------------------------------------------------
def _base_dir() -> Path:
    env = os.environ.get("PALETTENMINI_BERICHTE_DIR")
    if env:
        p = Path(env)
        p.mkdir(parents=True, exist_ok=True)
        return p
    if sys.platform == "win32":
        base = Path(os.environ.get("APPDATA", str(Path.home()))) / "PalettenMini"
    else:
        base = Path.home() / ".palettenmini"
    base.mkdir(parents=True, exist_ok=True)
    return base


def _berichte_dir() -> Path:
    d = _base_dir() / "berichte"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _archiv_pfad() -> Path:
    return _base_dir() / "berichte_archiv.json"


def archiv_dir_str() -> str:
    return str(_berichte_dir())


# ---------------------------------------------------------------------------
# Archiv
# ---------------------------------------------------------------------------
def _read_archiv() -> list[dict[str, Any]]:
    p = _archiv_pfad()
    if not p.exists():
        return []
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        return d if isinstance(d, list) else []
    except (OSError, json.JSONDecodeError):
        return []


def _write_archiv(eintraege: list[dict[str, Any]]) -> None:
    p = _archiv_pfad()
    tmp = p.with_suffix(p.suffix + ".tmp")
    tmp.write_text(json.dumps(eintraege, ensure_ascii=False, indent=2),
                    encoding="utf-8")
    tmp.replace(p)


def _dateiname(typ: str, ext: str) -> str:
    """Spec §7: youman_mini_<typ>_<YYYYMMDD_HHMM>.<ext>"""
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return f"youman_mini_{typ}_{ts}.{ext}"


def archiv_speichern(typ: str, dateiname: str, daten: bytes,
                       quell_lauf_id: str | None = None) -> dict[str, Any]:
    """Persistiert die Datei im Archiv-Ordner + ergänzt JSON-Index.
    Liefert dict mit id, pfad, size_kb."""
    pfad = _berichte_dir() / dateiname
    pfad.write_bytes(daten)
    eintrag = {
        "id": uuid.uuid4().hex,
        "typ": typ,
        "dateiname": dateiname,
        "pfad": str(pfad),
        "erzeugt_am": datetime.now().isoformat(timespec="seconds"),
        "quell_lauf_id": quell_lauf_id,
        "size_kb": round(len(daten) / 1024, 1),
    }
    h = _read_archiv()
    h.append(eintrag)
    _write_archiv(h)
    return eintrag


def archiv_liste(limit: int = 10) -> list[dict[str, Any]]:
    return sorted(_read_archiv(),
                   key=lambda e: e.get("erzeugt_am", ""),
                   reverse=True)[:limit]


def archiv_alle() -> list[dict[str, Any]]:
    return sorted(_read_archiv(),
                   key=lambda e: e.get("erzeugt_am", ""), reverse=True)


def archiv_lesen(eintrag_id: str) -> bytes | None:
    for e in _read_archiv():
        if e.get("id") == eintrag_id:
            try:
                return Path(e["pfad"]).read_bytes()
            except OSError:
                return None
    return None


def archiv_loeschen(eintrag_id: str) -> bool:
    h = _read_archiv()
    neu = []
    geloescht = False
    for e in h:
        if e.get("id") == eintrag_id:
            try:
                Path(e["pfad"]).unlink(missing_ok=True)
            except OSError:
                pass
            geloescht = True
            continue
        neu.append(e)
    if geloescht:
        _write_archiv(neu)
    return geloescht


# ---------------------------------------------------------------------------
# PDF-Helper (fpdf2 mit Latin-1 Encoding)
# ---------------------------------------------------------------------------
def _latin1(s: Any) -> str:
    """fpdf2 Core-Fonts unterstuetzen Latin-1. Unicode-Chars werden
    durch Aequivalente ersetzt."""
    if s is None:
        return ""
    text = str(s)
    ersatz = {
        "→": "->", "←": "<-", "↗": "/", "↘": "\\", "·": "*",
        "€": "EUR", "²": "2", "³": "3", "×": "x",
        "🟢": "[OK]", "🟡": "[!]", "🔴": "[X]", "⚫": "[?]",
        "✓": "OK", "✗": "X", "⚠": "!", "❌": "X", "✅": "OK",
        "📦": "", "📊": "", "📋": "", "💰": "", "🚚": "",
        "📄": "", "📂": "", "🎯": "", "⚖": "", "📈": "",
    }
    for k, v in ersatz.items():
        text = text.replace(k, v)
    # Final fallback: encode → decode mit replace
    return text.encode("latin-1", "replace").decode("latin-1")


class YoumanPDF(FPDF):
    """Standard-Layout mit Briefkopf + Footer."""

    def __init__(self, titel: str = "Youman Report"):
        super().__init__(format="A4")
        self.titel_text = titel
        self.set_margins(15, 18, 15)
        self.set_auto_page_break(auto=True, margin=18)

    def header(self):
        # Briefkopf
        self.set_font("Helvetica", "B", 16)
        self.set_text_color(15, 31, 61)
        self.cell(0, 8, _latin1("Youman Automation"), ln=True)
        self.set_font("Helvetica", "", 9)
        self.set_text_color(100, 100, 100)
        self.cell(0, 4, _latin1("Industriepaletten - Standardisierung"),
                   ln=True)
        # Trennlinie
        self.set_draw_color(15, 31, 61)
        self.set_line_width(0.5)
        y = self.get_y() + 2
        self.line(15, y, 195, y)
        # Bericht-Titel
        self.set_y(y + 3)
        self.set_font("Helvetica", "B", 14)
        self.set_text_color(15, 31, 61)
        self.cell(0, 8, _latin1(self.titel_text), ln=True)
        self.set_text_color(0, 0, 0)
        self.ln(2)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(150, 150, 150)
        ts = datetime.now().strftime("%d.%m.%Y %H:%M")
        try:
            from _build_info import BUILD_SHA, BUILD_VERSION
            stamp = f"Build {BUILD_SHA} / v{BUILD_VERSION}"
        except ImportError:
            stamp = "Build dev"
        self.cell(0, 6,
                   _latin1(f"Youman {stamp} | erzeugt {ts} | "
                            f"Seite {self.page_no()}/{{nb}}"),
                   align="C")


def _kpi_block(pdf: YoumanPDF, kpis: list[tuple[str, str]]) -> None:
    pdf.set_font("Helvetica", "B", 10)
    spaltenbreite = 60
    for label, wert in kpis:
        pdf.set_text_color(100, 100, 100)
        pdf.cell(spaltenbreite, 5, _latin1(label))
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(15, 31, 61)
    for label, wert in kpis:
        pdf.cell(spaltenbreite, 8, _latin1(wert))
    pdf.ln(10)
    pdf.set_text_color(0, 0, 0)


def _table(pdf: YoumanPDF, header: list[str], rows: list[list[str]],
            col_widths: list[int] | None = None) -> None:
    n = len(header)
    if col_widths is None:
        verfuegbar = 180
        col_widths = [verfuegbar // n] * n
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(230, 230, 230)
    for h, w in zip(header, col_widths):
        pdf.cell(w, 7, _latin1(h), border=1, fill=True, align="L")
    pdf.ln(7)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_fill_color(245, 245, 245)
    for i, row in enumerate(rows):
        if pdf.get_y() + 6 > pdf.h - 25:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_fill_color(230, 230, 230)
            for h, w in zip(header, col_widths):
                pdf.cell(w, 7, _latin1(h), border=1, fill=True, align="L")
            pdf.ln(7)
            pdf.set_font("Helvetica", "", 9)
            pdf.set_fill_color(245, 245, 245)
        fill = (i % 2 == 0)
        for v, w in zip(row, col_widths):
            pdf.cell(w, 6, _latin1(str(v))[:50], border=1, fill=fill)
        pdf.ln(6)


# ---------------------------------------------------------------------------
# Bericht 1: PDF Optimierungs-Report (§2)
# ---------------------------------------------------------------------------
def pdf_optimierungs_report(eintrag: dict, ergebnis: dict | None = None) -> bytes:
    """eintrag = ein Verlauf-Eintrag (mit optimierung-Snapshot).
    ergebnis = optional aktuelles ergebnis-Dict aus session (fuer Detail-
    Zuordnung). Wenn nicht angegeben, nutzt der Report nur den
    Verlauf-Snapshot."""
    opt = eintrag.get("optimierung", {}) or {}
    snap_p = opt.get("params_snapshot", {}) or {}
    snap_e = opt.get("ergebnis_snapshot", {}) or {}

    pdf = YoumanPDF("Optimierungs-Report")
    pdf.alias_nb_pages()
    pdf.add_page()

    # Stammdaten
    pdf.set_font("Helvetica", "", 10)
    datei = eintrag.get("datei_name", "—")
    datum = (eintrag.get("datum", "") or "")[:16].replace("T", " ")
    pdf.cell(0, 5, _latin1(f"Datei: {datei}"), ln=True)
    pdf.cell(0, 5, _latin1(f"Datum: {datum}"), ln=True)
    pdf.ln(2)

    # Parameter-Block
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(15, 31, 61)
    pdf.cell(0, 6, _latin1("Parameter"), ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 9)
    tol_str = (
        f"Toleranz: {opt.get('tol_kurz_pct', '?'):.0f}% / "
        f"{opt.get('tol_lang_pct', '?'):.0f}%"
        if opt.get("tol_modus") == "prozent"
        else f"Toleranz: {opt.get('tol_kurz_mm', '?')} / "
             f"{opt.get('tol_lang_mm', '?')} mm"
    )
    pdf.cell(0, 5, _latin1(tol_str), ln=True)
    pdf.cell(0, 5,
              _latin1(f"Kombinieren: {'an' if opt.get('kombinieren') else 'aus'} | "
                       f"Sonderpaletten: "
                       f"{'erlaubt' if snap_p.get('sonder_erlaubt', True) else 'verboten'} | "
                       f"Sonder-Deckel: {opt.get('sonder_deckel', 'frei')}"),
              ln=True)
    pdf.ln(3)

    # KPIs
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(15, 31, 61)
    pdf.cell(0, 6, _latin1("Kennzahlen"), ln=True)
    pdf.set_text_color(0, 0, 0)
    _kpi_block(pdf, [
        ("Standards", str(opt.get("standards", 0))),
        ("Sonder", str(opt.get("sonder", 0))),
        ("Gesamt", str(opt.get("gesamt", 0))),
    ])

    # Standards-Detail
    standards = snap_e.get("standards", [])
    if standards:
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(15, 31, 61)
        pdf.cell(0, 6, _latin1("Standardpaletten"), ln=True)
        pdf.set_text_color(0, 0, 0)
        rows = [[i + 1, f"{max(s)} x {min(s)} mm"]
                for i, s in enumerate(standards)]
        _table(pdf, ["#", "Massnahme (mm)"], rows, [20, 100])
        pdf.ln(3)

    # Zuordnungstabelle (wenn aktuelles Ergebnis vorhanden)
    if ergebnis and ergebnis.get("zuordnung"):
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(15, 31, 61)
        pdf.cell(0, 6, _latin1("Detail-Zuordnung"), ln=True)
        pdf.set_text_color(0, 0, 0)
        rows = []
        for z in ergebnis["zuordnung"][:200]:  # max 200 Zeilen
            rows.append([
                z.get("auftrag", "")[:15],
                z.get("name", "")[:18],
                f"{int(z.get('L', 0))}x{int(z.get('B', 0))}",
                z.get("ziel", "—"),
                str(z.get("menge", 0)),
                z.get("typ", ""),
            ])
        _table(pdf,
                ["Auftrag", "Kunde", "Artikel (mm)", "Ziel", "Menge", "Typ"],
                rows, [30, 35, 30, 35, 20, 30])
        if len(ergebnis["zuordnung"]) > 200:
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(0, 5,
                      _latin1(f"... {len(ergebnis['zuordnung']) - 200} "
                               f"weitere Zeilen (CSV-Export im Ergebnisse-Tab)."),
                      ln=True)

    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Bericht 2: Excel Zuordnung (§3)
# ---------------------------------------------------------------------------
def excel_zuordnung(eintrag: dict, ergebnis: dict | None = None) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    opt = eintrag.get("optimierung", {}) or {}

    # Sheet 1: Übersicht
    ws_u = wb.active
    ws_u.title = "Uebersicht"
    header_fill = PatternFill("solid", fgColor="E0E0E0")
    bold = Font(bold=True)
    ws_u["A1"] = "Optimierungs-Übersicht"
    ws_u["A1"].font = Font(bold=True, size=14, color="0F1F3D")
    ws_u["A3"] = "Datei:"
    ws_u["B3"] = eintrag.get("datei_name", "")
    ws_u["A4"] = "Datum:"
    ws_u["B4"] = (eintrag.get("datum", "") or "")[:16].replace("T", " ")
    ws_u["A6"] = "Parameter"
    ws_u["A6"].font = bold
    ws_u["A7"] = "Tol kurz (mm)"
    ws_u["B7"] = opt.get("tol_kurz_mm", "")
    ws_u["A8"] = "Tol lang (mm)"
    ws_u["B8"] = opt.get("tol_lang_mm", "")
    ws_u["A9"] = "Kombinieren"
    ws_u["B9"] = "ja" if opt.get("kombinieren") else "nein"
    ws_u["A10"] = "Sonder-Deckel"
    ws_u["B10"] = opt.get("sonder_deckel", "frei")
    ws_u["A12"] = "Kennzahlen"
    ws_u["A12"].font = bold
    ws_u["A13"] = "Standards"
    ws_u["B13"] = opt.get("standards", 0)
    ws_u["A14"] = "Sonder"
    ws_u["B14"] = opt.get("sonder", 0)
    ws_u["A15"] = "Gesamt"
    ws_u["B15"] = opt.get("gesamt", 0)
    for c in ("A6", "A12"):
        ws_u[c].fill = header_fill
    ws_u.column_dimensions["A"].width = 22
    ws_u.column_dimensions["B"].width = 25

    # Sheet 2: Zuordnung
    ws_z = wb.create_sheet("Zuordnung")
    ws_z.append([
        "Auftrag", "Kunde", "Artikelnummer", "Verbrauchsdatum",
        "Artikel L (mm)", "Artikel B (mm)", "Menge",
        "Neue Palette", "Typ", "Bemerkung",
    ])
    for c_ in ws_z[1]:
        c_.font = bold
        c_.fill = header_fill
    if ergebnis and ergebnis.get("zuordnung"):
        for z in ergebnis["zuordnung"]:
            bemerkung = z.get("begruendung", "")
            if z.get("typ") == "Standard":
                bemerkung = bemerkung or "innerhalb Toleranz"
            ws_z.append([
                z.get("auftrag", ""), z.get("name", ""),
                z.get("artikelnummer", ""), z.get("verbrauchsdatum", ""),
                int(z.get("L", 0)), int(z.get("B", 0)),
                int(z.get("menge", 0)),
                z.get("ziel", ""), z.get("typ", ""), bemerkung,
            ])
    # Auto-Filter + Spaltenbreiten
    ws_z.auto_filter.ref = ws_z.dimensions
    for col_idx, w in enumerate([12, 25, 18, 14, 13, 13, 8, 18, 18, 30], 1):
        ws_z.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    # Zahlen rechtsbündig
    for row in ws_z.iter_rows(min_row=2, min_col=5, max_col=7):
        for c_ in row:
            c_.alignment = Alignment(horizontal="right")

    # Sheet 3: Standards
    ws_s = wb.create_sheet("Standards")
    ws_s.append(["#", "Maß (mm)", "L (mm)", "B (mm)"])
    for c_ in ws_s[1]:
        c_.font = bold
        c_.fill = header_fill
    snap_e = opt.get("ergebnis_snapshot", {}) or {}
    for i, s in enumerate(snap_e.get("standards", []), 1):
        cs, cl = min(s), max(s)
        ws_s.append([i, f"{cl}x{cs}", cl, cs])
    ws_s.column_dimensions["A"].width = 6
    ws_s.column_dimensions["B"].width = 15

    # Sheet 4: Nicht zuordenbar
    ws_n = wb.create_sheet("Nicht_zuordenbar")
    ws_n.append(["Auftrag", "Kunde", "L", "B", "Menge"])
    for c_ in ws_n[1]:
        c_.font = bold
        c_.fill = header_fill
    if ergebnis:
        for nz in ergebnis.get("nicht_zuordenbar", []):
            ws_n.append([
                nz.get("auftrag", ""), nz.get("name", ""),
                int(nz.get("L", 0)), int(nz.get("B", 0)),
                int(nz.get("menge", 0)),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Bericht 3: Bestandsbericht (§4)
# ---------------------------------------------------------------------------
def bestandsbericht_pdf(eintraege: list[dict],
                         nur_bestand_positiv: bool = False,
                         nur_vollstaendig: bool = False) -> bytes:
    rows = []
    sum_bestand = 0
    sum_wert = 0.0
    for e in eintraege:
        bestand = int(e.get("bestand", 0) or 0)
        ek_lief = float(e.get("ek_lieferant_eur",
                                  e.get("einkaufspreis_eur", 0)) or 0)
        eigen = (float(e.get("selbstfertigung_material_eur", 0) or 0)
                  + float(e.get("selbstfertigung_lohn_eur", 0) or 0))
        if nur_bestand_positiv and bestand <= 0:
            continue
        if nur_vollstaendig and (ek_lief <= 0 or eigen <= 0):
            continue
        diff = ek_lief - eigen  # positiv = Eigenfertigung billiger
        wert = bestand * ek_lief
        sum_bestand += bestand
        sum_wert += wert
        rows.append([
            e.get("name", "")[:18] or "-",
            f"{e.get('L_mm', 0)} x {e.get('B_mm', 0)}",
            str(bestand),
            f"{ek_lief:.2f}",
            f"{eigen:.2f}",
            f"{diff:+.2f}",
            f"{wert:.2f}",
        ])

    pdf = YoumanPDF("Bestandsbericht")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5,
              _latin1(f"Stichtag: {datetime.now().strftime('%d.%m.%Y %H:%M')}"),
              ln=True)
    pdf.cell(0, 5, _latin1(f"Anzahl Typen: {len(rows)}"), ln=True)
    pdf.cell(0, 5, _latin1(f"Summe Bestand: {sum_bestand} Stk"), ln=True)
    pdf.cell(0, 5,
              _latin1(f"Inventarwert (EK Lieferant): {sum_wert:.2f} EUR"),
              ln=True)
    if nur_bestand_positiv or nur_vollstaendig:
        filter_text = []
        if nur_bestand_positiv:
            filter_text.append("Bestand > 0")
        if nur_vollstaendig:
            filter_text.append("Bezugskosten vollstaendig")
        pdf.cell(0, 5, _latin1(f"Filter: {' & '.join(filter_text)}"), ln=True)
    pdf.ln(3)
    _table(pdf,
            ["Name", "Masse", "Bestand", "EK Lief.", "Eigenfert.",
              "Delta", "S-Wert"],
            rows, [40, 28, 18, 18, 22, 18, 22])
    return bytes(pdf.output())


def bestandsbericht_excel(eintraege: list[dict],
                            nur_bestand_positiv: bool = False,
                            nur_vollstaendig: bool = False) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bestand"
    ws.append([
        "Name", "L (mm)", "B (mm)", "Bestand", "in Anlief.",
        "EK Lieferant", "Eigenfertigung", "Δ Lief-Eig",
        "Inventarwert", "Typ",
    ])
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="E0E0E0")
    for c_ in ws[1]:
        c_.font = bold; c_.fill = fill
    sum_b = 0; sum_w = 0.0
    for e in eintraege:
        bestand = int(e.get("bestand", 0) or 0)
        ek_lief = float(e.get("ek_lieferant_eur",
                                  e.get("einkaufspreis_eur", 0)) or 0)
        eigen = (float(e.get("selbstfertigung_material_eur", 0) or 0)
                  + float(e.get("selbstfertigung_lohn_eur", 0) or 0))
        if nur_bestand_positiv and bestand <= 0:
            continue
        if nur_vollstaendig and (ek_lief <= 0 or eigen <= 0):
            continue
        wert = bestand * ek_lief
        sum_b += bestand; sum_w += wert
        ws.append([
            e.get("name", ""), e.get("L_mm", 0), e.get("B_mm", 0),
            bestand, int(e.get("bestand_bestellt", 0) or 0),
            ek_lief, eigen, ek_lief - eigen, wert,
            e.get("typ", "standard"),
        ])
    # Aggregat
    n_last = ws.max_row + 2
    ws.cell(row=n_last, column=1, value="SUMME").font = bold
    ws.cell(row=n_last, column=4, value=sum_b).font = bold
    ws.cell(row=n_last, column=9, value=sum_w).font = bold
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    widths = [22, 10, 10, 10, 12, 10, 10, 12, 15, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=9):
        for c_ in row:
            c_.alignment = Alignment(horizontal="right")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Bericht 4: Wirtschaftlichkeitsbericht (§5)
# ---------------------------------------------------------------------------
def _trend_chart_png(daten: list[tuple[str, int, int]]) -> bytes | None:
    """Erzeugt ein einfaches Trend-Liniendiagramm (Standards in blau,
    Sonder in rot) als PNG via PIL — kein matplotlib-Dep."""
    if not daten or len(daten) < 2:
        return None
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return None
    W, H = 900, 280
    img = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(img)
    # Skala
    max_y = max(max(s for _, s, _ in daten), max(o for _, _, o in daten), 1)
    pad_l, pad_b, pad_t, pad_r = 50, 40, 20, 20
    plot_w = W - pad_l - pad_r
    plot_h = H - pad_t - pad_b
    # Achsen
    d.line([(pad_l, pad_t), (pad_l, H - pad_b)], fill="#94a3b8", width=1)
    d.line([(pad_l, H - pad_b), (W - pad_r, H - pad_b)], fill="#94a3b8", width=1)
    # Gridlines + Y-Labels
    try:
        font = ImageFont.load_default()
    except OSError:
        font = None
    for i in range(0, 5):
        y = H - pad_b - int(i * plot_h / 4)
        d.line([(pad_l, y), (W - pad_r, y)], fill="#f1f5f9", width=1)
        val = int(i * max_y / 4)
        if font:
            d.text((5, y - 6), str(val), fill="#64748b", font=font)
    # Datenpunkte
    n = len(daten)
    def x_at(i): return pad_l + int(i * plot_w / max(1, n - 1))
    def y_at(v): return H - pad_b - int(v * plot_h / max_y)
    pts_std = [(x_at(i), y_at(s)) for i, (_, s, _) in enumerate(daten)]
    pts_son = [(x_at(i), y_at(o)) for i, (_, _, o) in enumerate(daten)]
    if len(pts_std) >= 2:
        d.line(pts_std, fill="#2563eb", width=2)
    if len(pts_son) >= 2:
        d.line(pts_son, fill="#dc2626", width=2)
    # Punkte
    for x, y in pts_std:
        d.ellipse([(x - 3, y - 3), (x + 3, y + 3)], fill="#2563eb")
    for x, y in pts_son:
        d.ellipse([(x - 3, y - 3), (x + 3, y + 3)], fill="#dc2626")
    # Legende
    if font:
        d.rectangle([(W - 150, 10), (W - 138, 22)], fill="#2563eb")
        d.text((W - 132, 10), "Standards", fill="#1e293b", font=font)
        d.rectangle([(W - 70, 10), (W - 58, 22)], fill="#dc2626")
        d.text((W - 52, 10), "Sonder", fill="#1e293b", font=font)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def wirtschaftlichkeitsbericht_pdf(
    optimierungen: list[dict],
    zeitraum_label: str = "alle",
) -> bytes:
    pdf = YoumanPDF("Wirtschaftlichkeitsbericht")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5,
              _latin1(f"Zeitraum: {zeitraum_label} | "
                       f"{len(optimierungen)} Optimierungslaeufe"),
              ln=True)
    pdf.cell(0, 5,
              _latin1(f"Stichtag: {datetime.now().strftime('%d.%m.%Y %H:%M')}"),
              ln=True)
    pdf.ln(3)

    if not optimierungen:
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, _latin1("Keine Optimierungslaeufe im Zeitraum."),
                  ln=True)
        return bytes(pdf.output())

    # Aggregate
    avg_std = sum(int(e["optimierung"].get("standards", 0))
                    for e in optimierungen) / len(optimierungen)
    avg_son = sum(int(e["optimierung"].get("sonder", 0))
                    for e in optimierungen) / len(optimierungen)
    min_std = min(int(e["optimierung"].get("standards", 0))
                    for e in optimierungen)
    max_std = max(int(e["optimierung"].get("standards", 0))
                    for e in optimierungen)

    # Trend
    if len(optimierungen) >= 2:
        first_std = int(optimierungen[-1]["optimierung"].get("standards", 0))
        last_std = int(optimierungen[0]["optimierung"].get("standards", 0))
        trend = (("steigend" if last_std > first_std + 0.5 else
                  "fallend" if last_std < first_std - 0.5 else "stabil"))
    else:
        trend = "n/a"

    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(15, 31, 61)
    pdf.cell(0, 6, _latin1("Kennzahlen"), ln=True)
    pdf.set_text_color(0, 0, 0)
    _kpi_block(pdf, [
        ("Avg Standards", f"{avg_std:.1f}"),
        ("Avg Sonder", f"{avg_son:.1f}"),
        ("Min - Max Std", f"{min_std} - {max_std}"),
    ])

    # Auto-Management-Text
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(15, 31, 61)
    pdf.cell(0, 6, _latin1("Management-Zusammenfassung"), ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 5, _latin1(
        f"Im Berichtszeitraum wurden im Schnitt {avg_std:.1f} Standards "
        f"und {avg_son:.1f} Sonderpaletten pro Lauf bestimmt. "
        f"Die Anzahl der Standardpaletten ist {trend}. "
        f"{'Variantenreduktion wirkt - weiter so.' if trend == 'fallend' else ''}"
        f"{'Toleranz-Optimierung pruefen.' if trend == 'steigend' else ''}"
    ))
    pdf.ln(3)

    # Trend-Chart
    daten = []
    for e in reversed(optimierungen):  # chronologisch
        datum = (e.get("datum", "") or "")[:10]
        opt = e["optimierung"]
        daten.append((datum, int(opt.get("standards", 0)),
                       int(opt.get("sonder", 0))))
    chart_png = _trend_chart_png(daten)
    if chart_png:
        tmp = _berichte_dir() / f"_chart_{uuid.uuid4().hex[:8]}.png"
        tmp.write_bytes(chart_png)
        try:
            pdf.image(str(tmp), x=15, y=pdf.get_y(), w=180)
            pdf.ln(70)
        finally:
            tmp.unlink(missing_ok=True)

    # Vergleichstabelle
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(15, 31, 61)
    pdf.cell(0, 6, _latin1("Optimierungen im Zeitraum"), ln=True)
    pdf.set_text_color(0, 0, 0)
    rows = []
    for e in optimierungen[:30]:
        opt = e["optimierung"]
        rows.append([
            (e.get("datum", "") or "")[:10],
            (e.get("datei_name", "") or "")[:20],
            f"{opt.get('tol_kurz_mm', '?')}/{opt.get('tol_lang_mm', '?')}",
            str(opt.get("standards", 0)),
            str(opt.get("sonder", 0)),
            str(opt.get("gesamt", 0)),
        ])
    _table(pdf,
            ["Datum", "Datei", "Tol kurz/lang", "Std", "Son", "Ges"],
            rows, [25, 45, 30, 22, 22, 22])
    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Bericht 5: Lieferanten-Bestelldokument (§6)
# ---------------------------------------------------------------------------
def lieferanten_bestelldokument_pdf(
    lieferant: dict,
    positionen: list[dict],
    bestellnummer: str,
    bestelldatum: str = "",
    wunschtermin: str = "",
    bemerkung: str = "",
) -> bytes:
    """positionen = [{name, L, B, menge, ek_pro_stk_eur}, ...]"""
    pdf = YoumanPDF(f"Bestellung {bestellnummer}")
    pdf.alias_nb_pages()
    pdf.add_page()

    # Lieferanten-Adresse oben links
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, _latin1(f"An: {lieferant.get('name', '—')}"), ln=True)
    if lieferant.get("kontakt"):
        pdf.cell(0, 5, _latin1(lieferant["kontakt"]), ln=True)
    pdf.ln(3)

    # Meta-Block
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(40, 5, _latin1("Bestellnummer:"))
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, _latin1(bestellnummer), ln=True)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(40, 5, _latin1("Bestelldatum:"))
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5,
              _latin1(bestelldatum or datetime.now().strftime("%d.%m.%Y")),
              ln=True)
    if wunschtermin:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(40, 5, _latin1("Wunschtermin:"))
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 5, _latin1(wunschtermin), ln=True)
    pdf.ln(5)

    # Positionstabelle
    rows, summe = [], 0.0
    for i, p in enumerate(positionen, 1):
        m = int(p.get("menge", 0))
        ek = float(p.get("ek_pro_stk_eur", 0))
        s = m * ek
        summe += s
        rows.append([
            str(i),
            p.get("name", "")[:20] or "Palette",
            f"{int(p.get('L', 0))} x {int(p.get('B', 0))}",
            str(m),
            f"{ek:.2f}",
            f"{s:.2f}",
        ])
    _table(pdf, ["Pos.", "Palettentyp", "Masse", "Menge", "EK", "Summe"],
            rows, [15, 60, 30, 22, 25, 30])
    pdf.ln(3)

    # Summen
    pdf.set_font("Helvetica", "", 10)
    netto = summe
    mwst = netto * 0.19
    brutto = netto + mwst
    pdf.cell(120)
    pdf.cell(35, 5, _latin1("Summe netto:"))
    pdf.cell(0, 5, _latin1(f"{netto:.2f} EUR"), align="R", ln=True)
    pdf.cell(120)
    pdf.cell(35, 5, _latin1("zzgl. 19% USt:"))
    pdf.cell(0, 5, _latin1(f"{mwst:.2f} EUR"), align="R", ln=True)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(120)
    pdf.cell(35, 5, _latin1("Summe brutto:"))
    pdf.cell(0, 5, _latin1(f"{brutto:.2f} EUR"), align="R", ln=True)

    if bemerkung:
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 5, _latin1("Bemerkung:"), ln=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(0, 5, _latin1(bemerkung))

    pdf.ln(6)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.multi_cell(0, 4, _latin1(
        "Es gelten die AGB des Bestellers. Lieferung frei Haus, "
        "Zahlungsziel 14 Tage netto, soweit nicht anders vereinbart."
    ))
    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Bericht 6: Wiederbeschaffungszeiten (Spec §5)
# ---------------------------------------------------------------------------
def wiederbeschaffungs_pdf(rows: list[dict],
                             sicherheitspuffer_tage: int = 7) -> bytes:
    """rows = aus wiederbeschaffung.baue_zeilen()."""
    pdf = YoumanPDF("Wiederbeschaffungszeiten")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5,
              _latin1(f"Stichtag: {datetime.now().strftime('%d.%m.%Y %H:%M')} "
                       f"| Sicherheitspuffer: {sicherheitspuffer_tage} Tage"),
              ln=True)
    krit = [r for r in rows if r["kritisch"]]
    pdf.cell(0, 5,
              _latin1(f"Paletten gesamt: {len(rows)} "
                       f"| davon kritisch: {len(krit)}"),
              ln=True)
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(0, 5, _latin1(
        "Ampel: [OK] Reichweite > Lieferzeit + Puffer | "
        "[!] Reichweite ~ Lieferzeit | "
        "[X] Reichweite < Lieferzeit (Engpass) | "
        "[?] ohne Verbrauchsdaten"
    ))
    pdf.ln(2)

    table_rows = []
    for r in rows:
        rw = (f"{r['reichweite_tage']:.0f}"
              if r['reichweite_tage'] is not None else "—")
        table_rows.append([
            (r["name"] or "—")[:18],
            f"{r['L_mm']}x{r['B_mm']}",
            (r["lieferant"] or "—")[:18],
            str(r["lieferzeit_tage"]),
            str(r["bestand"]),
            str(r["bestand_bestellt"]),
            rw,
            r["ampel"],
        ])
    _table(pdf,
            ["Palette", "Masse (mm)", "Lieferant", "LZ (T)",
              "Bestand", "in Anl.", "Reichw. (T)", "Status"],
            table_rows, [30, 22, 30, 16, 20, 18, 22, 22])
    return bytes(pdf.output())


def wiederbeschaffungs_excel(rows: list[dict]) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wiederbeschaffung"
    ws.append(["Palette", "L", "B", "Lieferant", "Lieferzeit (T)",
                "Bestand", "in Anlieferung", "Reichweite (T)", "Ampel"])
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="E0E0E0")
    for c_ in ws[1]:
        c_.font = bold; c_.fill = fill
    for r in rows:
        ws.append([
            r["name"] or "—", r["L_mm"], r["B_mm"],
            r["lieferant"] or "—", r["lieferzeit_tage"],
            r["bestand"], r["bestand_bestellt"],
            r["reichweite_tage"] if r["reichweite_tage"] is not None else "—",
            r["ampel"],
        ])
    ws.auto_filter.ref = ws.dimensions
    for col_idx, w in enumerate([22, 8, 8, 22, 12, 10, 14, 14, 8], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=8):
        for c_ in row:
            c_.alignment = Alignment(horizontal="right")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Bericht 7: Preisentwicklung (Spec §8)
# ---------------------------------------------------------------------------
def preisentwicklung_pdf(zeitreihe: dict[str, list[dict]]) -> bytes:
    """zeitreihe = aus preis_historie.preis_zeitreihe()."""
    pdf = YoumanPDF("Preisentwicklung")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5,
              _latin1(f"Stichtag: {datetime.now().strftime('%d.%m.%Y %H:%M')}"),
              ln=True)
    pdf.cell(0, 5, _latin1(f"Maße mit Preis-Historie: {len(zeitreihe)}"),
              ln=True)
    pdf.ln(3)

    if not zeitreihe:
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, _latin1(
            "Keine Preis-Snapshots im Verlauf gespeichert. Erst nach "
            "mehreren Importen entsteht eine Historie."
        ), ln=True)
        return bytes(pdf.output())

    # Pro Maß: Tabelle + kleines Trend-Linien-Mini-Chart
    for mass, punkte in sorted(zeitreihe.items()):
        if len(punkte) < 1:
            continue
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(15, 31, 61)
        pdf.cell(0, 6, _latin1(f"Palette {mass}"), ln=True)
        pdf.set_text_color(0, 0, 0)
        rows = []
        for p in punkte:
            rows.append([p["datum"], f"{p['ek']:.2f}"])
        if len(punkte) >= 2:
            erst = punkte[0]["ek"]; letzt = punkte[-1]["ek"]
            diff = letzt - erst
            pct = (diff / erst * 100) if erst > 0 else 0
            pfeil = "->" if abs(pct) < 0.5 else ("/" if diff > 0 else "\\")
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(0, 5, _latin1(
                f"Trend: {erst:.2f} EUR -> {letzt:.2f} EUR  "
                f"({pfeil} {pct:+.1f} %)"), ln=True)
        _table(pdf, ["Datum", "EK (EUR)"], rows, [40, 30])
        pdf.ln(3)
    return bytes(pdf.output())

