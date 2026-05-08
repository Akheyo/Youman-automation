"""Konvertiert die Draht-Müller-Auftragsliste (TSV) in das App-Excel-Format.

Eingabe: ``data/draht_mueller_raw.tsv`` mit Spalten
    Maschine | KW Fertigung | Nr. | KW Lief. | Auftrag | ANr | AK | Name |
    (leer) | Artikelnummer | (leer) | Masche/Bemerkung | Menge | Stck Pal. |
    P-Länge | P-Breite | P-Höhe | P-Anzahl

Ausgabe: ``data/beispiel_palettenliste.xlsx`` mit Spalten
    Artikelnummer | Laenge | Breite | Hoehe | Benoetigte_Paletten |
    Stueckzahl_pro_Palette | Palettenkosten | Kunde | Auftrag | KW_Lieferung
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def parse_tsv(pfad: Path) -> list[dict]:
    rows: list[dict] = []
    with pfad.open("r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter="\t")
        header_seen = False
        for parts in reader:
            if not parts:
                continue
            first = parts[0].strip() if parts else ""
            if not header_seen:
                if first.lower().startswith("maschine"):
                    header_seen = True
                continue
            if first != "Paletten":
                continue
            # Erwartet 18 Spalten (mit zwei Trenn-Leerspalten):
            # 0 Maschine | 1 KW Fert | 2 Nr | 3 KW Lief | 4 Auftrag | 5 ANr |
            # 6 AK | 7 Name | 8 (leer) | 9 Artikel | 10 (leer) | 11 Bemerkung |
            # 12 Menge | 13 Stck Pal | 14 P-Länge | 15 P-Breite | 16 P-Höhe |
            # 17 P-Anzahl
            while len(parts) < 18:
                parts.append("")

            def _f(idx: int) -> float:
                v = parts[idx].strip()
                if not v:
                    return 0.0
                try:
                    return float(v.replace(",", "."))
                except ValueError:
                    return 0.0

            def _i(idx: int) -> int:
                return int(_f(idx))

            laenge = _f(14)
            breite = _f(15)
            if laenge <= 0 or breite <= 0:
                continue
            stk_pal = _i(13)
            # 9999 = "individuelle Verpackung", für die Optimierung nicht
            # repräsentativ — Stk/Pal lassen wir aber stehen, da sie bekannt
            # sind. Nur tatsächlich fehlende Maße werden gefiltert.
            p_anzahl = _i(17) or 1
            artikel = parts[9].strip() or parts[10].strip()
            if not artikel:
                continue

            rows.append({
                "Artikelnummer": artikel,
                "Laenge": laenge,
                "Breite": breite,
                "Hoehe": _f(16),
                "Benoetigte_Paletten": p_anzahl,
                "Stueckzahl_pro_Palette": stk_pal,
                "Palettenkosten": 14.0,  # Default
                "Kunde": parts[7].strip(),
                "Auftrag": parts[4].strip(),
                "KW_Lieferung": parts[3].strip(),
                "Menge": _i(12),
                "KW_Fertigung": parts[1].strip(),
            })
    return rows


def schreibe_excel(rows: list[dict], ziel: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Palettenliste"

    spalten = [
        "Artikelnummer",
        "Laenge",
        "Breite",
        "Hoehe",
        "Benoetigte_Paletten",
        "Stueckzahl_pro_Palette",
        "Palettenkosten",
        "Kunde",
        "Auftrag",
        "KW_Lieferung",
        "Menge",
        "KW_Fertigung",
    ]
    ws.append(spalten)

    fill = PatternFill("solid", fgColor="1A2944")
    font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(spalten, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    for r in rows:
        ws.append([r.get(s, "") for s in spalten])

    ziel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ziel)


def main() -> None:
    base = Path(__file__).resolve().parent.parent
    quelle = base / "data" / "draht_mueller_raw.tsv"
    if not quelle.exists():
        print(f"Quelle nicht gefunden: {quelle}", file=sys.stderr)
        sys.exit(2)

    rows = parse_tsv(quelle)
    print(f"Geparste Zeilen: {len(rows)}")
    if not rows:
        print("Keine validen Zeilen gefunden.", file=sys.stderr)
        sys.exit(3)

    ziel_beispiel = base / "data" / "beispiel_palettenliste.xlsx"
    schreibe_excel(rows, ziel_beispiel)
    print(f"Beispieldatei: {ziel_beispiel} ({ziel_beispiel.stat().st_size} Bytes)")

    ziel_voll = base / "data" / "draht_mueller_palettenliste.xlsx"
    schreibe_excel(rows, ziel_voll)
    print(f"Vollständige Datei: {ziel_voll} ({ziel_voll.stat().st_size} Bytes)")


if __name__ == "__main__":
    main()
