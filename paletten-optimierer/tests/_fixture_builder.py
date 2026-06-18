"""Erzeugt zur Testzeit eine xlsx mit der EXAKTEN Draht-Mueller-
Struktur — Code-only, KEINE Abhaengigkeit von externen Test-Dateien.

Struktur:
- Zeile 1: Firmenname
- Zeile 2: leer
- Zeile 3: Metadaten (links)
- Zeile 4: leer
- Zeile 5: Header (oder konfigurierbar)
- Zeile 6: leer
- Daten ab Zeile (header_zeile+2), je eine Leerzeile dazwischen
- Spalten: E=Auftrag, H=Name, J=Artikelnummer, M=Menge, N=Stck Pal,
           O=P-Länge, P=P-Breite, Q=P-Höhe, R=P-Anzahl
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

import openpyxl


def erzeuge_fixture(rows: Iterable[dict],
                     ziel_pfad: Path,
                     header_zeile: int = 5) -> Path:
    """Baut eine .xlsx an ``ziel_pfad``.

    ``rows`` = Iterable von dicts mit Keys (alle optional):
        auftrag, name, artikelnr, menge, L (P-Länge), B (P-Breite),
        hoehe (P-Höhe), p_anzahl (Ablenkung, Default 1)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabelle1"

    # Header-Zeile
    if header_zeile >= 1:
        ws.cell(row=1, column=1, value="Test Müller GmbH & Co. KG")
    if header_zeile >= 3:
        ws.cell(row=3, column=1, value="verplante Maschinen")
        ws.cell(row=3, column=11, value="Datum '07.05.26' 20:53, Seite 1")

    h = header_zeile
    ws.cell(row=h, column=1,  value="Maschine")
    ws.cell(row=h, column=2,  value="KW Fertigung")
    ws.cell(row=h, column=3,  value="Nr.")
    ws.cell(row=h, column=4,  value="KW Lief")
    ws.cell(row=h, column=5,  value="Auftrag")
    ws.cell(row=h, column=6,  value="ANr")
    ws.cell(row=h, column=7,  value="AK")
    ws.cell(row=h, column=8,  value="Name")
    ws.cell(row=h, column=10, value="Artikelnummer")
    ws.cell(row=h, column=12, value="Masche/Bemerkung")
    ws.cell(row=h, column=13, value="Menge")
    ws.cell(row=h, column=14, value="Stck Pal")
    ws.cell(row=h, column=15, value="P-Länge")
    ws.cell(row=h, column=16, value="P-Breite")
    ws.cell(row=h, column=17, value="P-Höhe")
    ws.cell(row=h, column=18, value="P-Anzahl")

    aktuell = h + 2  # eine Leerzeile zwischen Header und ersten Daten
    for row in rows:
        ws.cell(row=aktuell, column=1,  value="Paletten")
        ws.cell(row=aktuell, column=5,  value=row.get("auftrag", ""))
        ws.cell(row=aktuell, column=8,  value=row.get("name", ""))
        ws.cell(row=aktuell, column=10, value=row.get("artikelnr", ""))
        ws.cell(row=aktuell, column=13, value=row.get("menge", 1))
        ws.cell(row=aktuell, column=14, value=row.get("stck_pal", 100))
        if row.get("L") is not None:
            ws.cell(row=aktuell, column=15, value=row["L"])
        if row.get("B") is not None:
            ws.cell(row=aktuell, column=16, value=row["B"])
        if row.get("hoehe") is not None:
            ws.cell(row=aktuell, column=17, value=row["hoehe"])
        ws.cell(row=aktuell, column=18, value=row.get("p_anzahl", 1))
        aktuell += 2  # Leerzeile nach jeder Datenzeile

    ziel_pfad.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ziel_pfad)
    return ziel_pfad
