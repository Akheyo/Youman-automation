"""End-to-End-Fixture fuer E2 (ANr) + E3 (Puffer) Tests.

Struktur:
- 3 Aufträge mit derselben AW aber verschiedenen ANrs (E2-Dedup-Test)
- Standardgrößen so gewählt, dass die Bedarfe im Ergebnis
  ≠ Puffer-Vielfache sind (E3-Delta sichtbar)

15 Datenzeilen, seed=99 fuer Determinismus.
"""
from __future__ import annotations

import random
from pathlib import Path

import openpyxl


HEADER = [
    "Maschine",        # A
    "KW Fertigung",    # B
    "Nr.",             # C
    "KW Lief",         # D
    "Auftrag",         # E
    "ANr",             # F  ← E2
    "AK",              # G
    "Name",            # H
    "Artikelnummer",   # I
    None,              # J leer
    None,              # K leer
    "Masche/Bemerkung",# L
    "Menge",           # M
    "Stck Pal",        # N
    "P-Länge",         # O
    "P-Breite",        # P
    "P-Höhe",          # Q
    "P-Anzahl",        # R
]


def _daten() -> list[tuple]:
    rng = random.Random(99)
    daten = []
    # 3 Aufträge mit gleicher AW aber verschiedenen ANrs — Draht Müller
    # Kunde ruft denselben Auftrag mehrfach ab.
    daten.append(("Paletten", 2701, 1, 2705, "200100", 1, "", "DM-Kunde A",
                   "ART-DM-01", None, None, "", 28, 300, 1250, 850, 1100, 1))
    daten.append(("Paletten", 2701, 2, 2705, "200100", 2, "", "DM-Kunde A",
                   "ART-DM-01", None, None, "", 30, 300, 1250, 850, 1100, 1))
    daten.append(("Paletten", 2701, 3, 2705, "200100", 3, "", "DM-Kunde A",
                   "ART-DM-01", None, None, "", 20, 300, 1250, 850, 1100, 1))
    # Weitere Aufträge mit Menge != Puffer-Vielfaches
    for i, (menge, l, b) in enumerate([
        (18, 850, 1250),  # gleiches Maß, andere AW
        (23, 850, 1250),
        (37, 850, 1250),
        (12, 1250, 850),  # gedreht
        (11, 800, 1200),  # zweites Cluster
        (14, 800, 1200),
        (17, 800, 1200),
        (7,  900, 1500),  # drittes Cluster
        (13, 900, 1500),
        (9,  900, 1500),
        (16, 900, 1500),
        (6,  900, 1500),
    ]):
        nr = 4 + i
        auftrag = 200100 + nr
        anr = rng.randint(1, 4)
        daten.append((
            "Paletten", 2701, nr, 2705, str(auftrag), anr, "",
            f"DM-Kunde {chr(65 + i % 5)}", f"ART-DM-{nr:02d}",
            None, None, "", menge, 300, l, b, 1100, 1,
        ))
    return daten


def baue(pfad: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="DM-E2E-Test Fixture")
    ws.cell(row=3, column=1, value="verplante Maschinen")
    ws.cell(row=3, column=11, value="Datum 04.07.26")
    for col, val in enumerate(HEADER, start=1):
        ws.cell(row=5, column=col, value=val)

    zeile = 7
    for daten_zeile in _daten():
        for col, val in enumerate(daten_zeile, start=1):
            ws.cell(row=zeile, column=col, value=val)
        zeile += 2

    pfad.parent.mkdir(parents=True, exist_ok=True)
    wb.save(pfad)
    return pfad


if __name__ == "__main__":
    ziel = Path(__file__).resolve().parent / "e2e_anr.xlsx"
    p = baue(ziel)
    print(f"Fixture geschrieben: {p} ({p.stat().st_size} Bytes)")
    print("15 Datenzeilen. AW 200100 hat 3 ANrs (1, 2, 3).")
    print("Bedarfsmengen sind absichtlich nicht Vielfache von 5, "
          "damit Puffer-Delta sichtbar wird.")
