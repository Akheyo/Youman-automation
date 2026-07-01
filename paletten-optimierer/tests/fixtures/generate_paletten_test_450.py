"""Deterministischer Generator fuer Paletten_Test_450.xlsx.

Struktur wie Draht-Mueller-Vorlage:
  Zeile 1: Firmenname
  Zeile 2: leer
  Zeile 3: Metadaten
  Zeile 4: leer
  Zeile 5: Header
  Zeile 6: leer
  Ab Zeile 7: Datenzeilen, LEERZEILE zwischen jeder Datenzeile
             (data_row += 2)

15 Cluster-Zentren in PRODUKT-Massen (Excel-Werte = Produkt + 50 in
beiden Achsen). 30 Auftraege pro Cluster mit +-25 mm Variation, 50%
gedreht, Menge 1..10, Hoehe zufaellig aus (1000,1200,1500).
Auftragsnummern ab 100000 aufsteigend, danach random.shuffle.

Nach Import + optimiere(tol_kurz=100, tol_lang=100, kombi=3, hetero=True,
sonder_deckel=5) muessen genau 15 Standards herauskommen.
"""
from __future__ import annotations

import random
from pathlib import Path

import openpyxl


CLUSTER_ZENTREN = [
    (380, 580),   (470, 770),   (560, 920),
    (650, 1080),  (760, 1250),  (850, 1420),
    (940, 1580),  (1030, 1740), (1140, 1890),
    (1230, 2050), (1320, 2200), (1430, 2360),
    (1520, 2510), (1610, 2670), (1720, 2820),
]
AUFTRAEGE_PRO_CLUSTER = 30
VARIATION_MM = 25
PALETTEN_AUFSCHLAG_MM = 50   # Excel-Wert = Produkt + Aufschlag
HOEHEN = (1000, 1200, 1500)

HEADER_ZEILE = 5
DATEN_START = 7

# Spalten wie in der echten Draht-Mueller-Datei (1-basiert)
HEADER = [
    "Maschine",         # A / 1
    "KW Fertigung",     # B / 2
    "Nr.",              # C / 3
    "KW Lief",          # D / 4
    "Auftrag",          # E / 5
    "ANr",              # F / 6
    "AK",               # G / 7
    "Name",             # H / 8
    "Artikelnummer",    # I / 9
    None,               # J / 10 (leer)
    None,               # K / 11 (leer)
    "Masche/Bemerkung", # L / 12
    "Menge",            # M / 13
    "Stck Pal",         # N / 14
    "P-Länge",          # O / 15
    "P-Breite",         # P / 16
    "P-Höhe",           # Q / 17
    "P-Anzahl",         # R / 18
]


def _erzeuge_orders(rng: random.Random) -> list[dict]:
    orders: list[dict] = []
    nr = 100000
    for cx, cy in CLUSTER_ZENTREN:
        for _ in range(AUFTRAEGE_PRO_CLUSTER):
            dx = rng.randint(-VARIATION_MM, VARIATION_MM)
            dy = rng.randint(-VARIATION_MM, VARIATION_MM)
            produkt_L = cy + dy  # laengere Achse
            produkt_B = cx + dx  # kuerzere Achse
            excel_L = produkt_L + PALETTEN_AUFSCHLAG_MM
            excel_B = produkt_B + PALETTEN_AUFSCHLAG_MM
            # 50 % gedreht — dann steht Kurz auf P-Laenge, Lang auf P-Breite
            if rng.random() < 0.5:
                excel_L, excel_B = excel_B, excel_L
            orders.append({
                "auftrag": str(nr),
                "name": f"Kunde-{nr % 50}",
                "artikelnummer": f"ART-{nr:06d}",
                "menge": rng.randint(1, 10),
                "stk_pal": 10,
                "p_laenge": excel_L,
                "p_breite": excel_B,
                "p_hoehe": rng.choice(HOEHEN),
            })
            nr += 1
    rng.shuffle(orders)
    return orders


def baue(pfad: Path) -> Path:
    rng = random.Random(42)
    orders = _erzeuge_orders(rng)
    assert len(orders) == 450, f"Erwartet 450, bekam {len(orders)}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabelle1"

    ws.cell(row=1, column=1, value="TEST-FIRMA GmbH — Testdaten")
    ws.cell(row=3, column=1, value="verplante Maschinen")
    ws.cell(row=3, column=11, value="Datum '01.07.26' 00:00, Seite 1")

    for col_idx, val in enumerate(HEADER, start=1):
        ws.cell(row=HEADER_ZEILE, column=col_idx, value=val)

    zeile = DATEN_START
    for i, o in enumerate(orders, start=1):
        ws.cell(row=zeile, column=1,  value="Paletten")
        ws.cell(row=zeile, column=2,  value=2626)
        ws.cell(row=zeile, column=3,  value=i)
        ws.cell(row=zeile, column=4,  value=2627)
        ws.cell(row=zeile, column=5,  value=o["auftrag"])
        ws.cell(row=zeile, column=6,  value=0)
        ws.cell(row=zeile, column=7,  value="")
        ws.cell(row=zeile, column=8,  value=o["name"])
        ws.cell(row=zeile, column=9,  value=o["artikelnummer"])
        # Spalten 10, 11 bleiben leer
        ws.cell(row=zeile, column=12, value="")
        ws.cell(row=zeile, column=13, value=o["menge"])
        ws.cell(row=zeile, column=14, value=o["stk_pal"])
        ws.cell(row=zeile, column=15, value=o["p_laenge"])
        ws.cell(row=zeile, column=16, value=o["p_breite"])
        ws.cell(row=zeile, column=17, value=o["p_hoehe"])
        ws.cell(row=zeile, column=18, value=o["menge"])
        zeile += 2   # LEERZEILE zwischen jeder Datenzeile

    pfad.parent.mkdir(parents=True, exist_ok=True)
    wb.save(pfad)
    return pfad


if __name__ == "__main__":
    ziel = Path(__file__).resolve().parent / "Paletten_Test_450.xlsx"
    p = baue(ziel)
    size = p.stat().st_size
    print(f"Fixture geschrieben: {p} ({size:,} Bytes)")
    print("15 Cluster x 30 Auftraege = 450 Datenzeilen, seed=42.")
