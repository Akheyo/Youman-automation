"""Erzeugt eine Test-Fixture mit der EXAKTEN Struktur der
Draht-Müller-Excel:

- Zeile 1: Firmenname ('Draht Müller GmbH & Co. KG')
- Zeile 2: leer
- Zeile 3: 'verplante Maschinen' (links) + Datum/Seite (rechts in Spalte K)
- Zeile 4: leer
- Zeile 5: Header
- Zeile 6: leer
- Zeile 7, 9, 11, ...: Daten — JEDE zweite Zeile ist Daten,
  dazwischen Leerzeilen.

Daten = die ersten 15 Aufträge mit Menge-Werten 2,2,1,5,48,1,1,2,
9,8,2,2,10,1,2 (Summe 96, GI-RO 109767 mit 48).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl


# Header-Reihenfolge wie in der echten Datei (Spalten A-R, mit J+K leer)
HEADER = [
    "Maschine",        # A
    "KW Fertigung",    # B
    "Nr.",             # C
    "KW Lief",         # D
    "Auftrag",         # E
    "ANr",             # F
    "AK",              # G
    "Name",            # H
    None,              # I leer
    "Artikelnummer",   # J  (eigentlich I laut Spec, hier korrigiert auf realer Datei)
    None,              # K leer
    "Masche/Bemerkung",  # L
    "Menge",           # M  ← Palettenanzahl
    "Stck Pal",        # N
    "P-Länge",         # O
    "P-Breite",        # P
    "P-Höhe",          # Q
    "P-Anzahl",        # R  ← wird ignoriert
]

# 15 Datenzeilen mit den exakten Mengen aus dem User-Spec
DATEN = [
    ("Paletten", 2619,  1,    0, 110135, 0, "", "VDL Jansen bv",        None, "PSCJ003637-0450 /001", None, "", 2, 350, 1220,  950, 1100, 1),
    ("Paletten", 2619,  2, 2620, 110287, 0, "", "BITO Lagertechnik",    None, "PSC2403x1508    /001", None, "", 2, 100, 2450, 1550, "",   1),
    ("Paletten", 2619,  3, 2620, 109931, 0, "", "Hellmann Poultry Gmb", None, "PSC45-36-0416   /004", None, "", 1, 100, 2420, 1050,  790, 1),
    ("Paletten", 2619,  4, 2623, 110021, 0, "", "Hellmann Poultry Gmb", None, "PSC45-36-0416   /004", None, "", 5, 100, 2420, 1050,  790, 1),
    ("Paletten", 2619,  5, 2622, 109767, 0, "06/06", "GI-RO Technik GmbH &", None, "EWM822x880      /001", None, "", 48, 38, 910, 850, 1300, 2),
    ("Paletten", 2619,  6, 2622, 110017, 0, "", "Hellmann Poultry Gmb", None, "PSC45-36-1049   /001", None, "", 1, 100, 2415,  760,  780, 1),
    ("Paletten", 2619,  7, 2621, 110150, 0, "", "R. Inauen AG",         None, "PSV1995x442     /001", None, "", 1, 100, 2050,  950, "",   1),
    ("Paletten", 2619,  8, 2621, 110203, 0, "", "Bressers Metaal BV",   None, "PSV1160x840     /001", None, "", 2, 125, 1200,  800,  950, 1),
    ("Paletten", 2619,  9, 2623, 110070, 0, "", "Big Dutchman",         None, "PSF83-50-4400   /001", None, "", 9, 112, 1620, 1225, 1340, 1),
    ("Paletten", 2619, 10,    0, 109971, 0, "", "VDL Jansen bv",        None, "PSCJ2.5X50X50ZNA/001", None, "", 8, 125, 3050, 1550,  500, 1),
    ("Paletten", 2619, 11, 2634, 110260, 0, "", "Hellmann Poultry Gmb", None, "PSC45-36-0868   /001", None, "", 2, 500, 1180,  800,  930, 1),
    ("Paletten", 2619, 12, 2633, 110224, 0, "", "Hellmann Poultry Gmb", None, "PSC45-36-0868   /001", None, "", 2, 500, 1180,  800,  930, 1),
    ("Paletten", 2619, 13, 2610, 109576, 0, "03/05", "Big Dutchman",    None, "PSF83-07-2880   /002", None, "", 10, 112, 1430, 1200, 1200, 1),
    ("Paletten", 2619, 14, 2621, 110055, 0, "", "Urban GmbH & Co.",     None, "PSB1631x441     /006", None, "", 1, 50, 1680,  500, "",   1),
    ("Paletten", 2619, 15, 2621, 110055, 0, "", "Urban GmbH & Co.",     None, "PSB1631x441     /004", None, "", 2, 50, 1680,  500, "",   1),
]


def baue_fixture(pfad: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabelle1"

    # Zeile 1: Firmenname
    ws.cell(row=1, column=1, value="Draht Müller GmbH & Co. KG")
    # Zeile 3: Metadata + Datum/Seite (Datum landet in Spalte K)
    ws.cell(row=3, column=1, value="verplante Maschinen")
    ws.cell(row=3, column=11, value="Datum '07.05.26' 20:53, Seite 1")

    # Zeile 5: Header
    for col_idx, val in enumerate(HEADER, start=1):
        ws.cell(row=5, column=col_idx, value=val)

    # Daten ab Zeile 7, mit Leerzeile vor jeder Datenzeile (außer 7)
    # Zeile 6 = leer, 7 = Daten, 8 = leer, 9 = Daten, ...
    current_row = 7
    for daten_zeile in DATEN:
        for col_idx, val in enumerate(daten_zeile, start=1):
            ws.cell(row=current_row, column=col_idx, value=val)
        current_row += 2  # Leerzeile dazwischen

    pfad.parent.mkdir(parents=True, exist_ok=True)
    wb.save(pfad)


if __name__ == "__main__":
    ziel = Path(__file__).resolve().parent / "draht_mueller_anker.xlsx"
    baue_fixture(ziel)
    print(f"Fixture erstellt: {ziel} ({ziel.stat().st_size} bytes)")
    print(f"15 Datenzeilen, jeweils mit Leerzeile dazwischen.")
    print(f"Menge-Werte: 2,2,1,5,48,1,1,2,9,8,2,2,10,1,2 (Summe 96)")
    print(f"GI-RO 109767 in Zeile 15 mit Menge=48")
