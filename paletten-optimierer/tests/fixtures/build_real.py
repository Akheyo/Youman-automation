"""Erzeugt tests/fixtures/draht_mueller_real.xlsx strikt nach
User-Spec.

Struktur:
- Z1, Spalte A: "Draht Müller GmbH & Co. KG"
- Z2: komplett leer
- Z3, Spalte A: "verplante Maschinen";
  Z3, Spalte K: "Datum '07.05.26' 20:53, Seite 1"
- Z4: komplett leer
- Z5 = Header:
    A=Maschine B=KW Fertigung C=Nr. D=KW Lief E=Auftrag F=ANr
    G=AK H=Name I=Artikelnummer L=Masche/Bemerkung M=Menge
    N=Stck Pal O=P-Länge P=P-Breite Q=P-Höhe R=P-Anzahl
  (J und K im Header bleiben leer)
- Z6: komplett leer
- Daten in ungeraden Zeilen ab Z7, jeweils mit Leerzeile dazwischen
  (Z8, Z10, Z12, ... bleiben leer). Genau 15 Datenzeilen.

Nur die vom User benannten Spalten werden in den Datenzeilen
gefüllt — alle anderen Spalten bleiben leer.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl


# 1-basierte Spalten-Indizes
COL_AUFTRAG       = 5   # E
COL_NAME          = 8   # H
COL_ARTIKELNUMMER = 9   # I
COL_MENGE         = 13  # M
COL_STCK_PAL      = 14  # N
COL_LAENGE        = 15  # O
COL_BREITE        = 16  # P
COL_HOEHE         = 17  # Q
COL_P_ANZAHL      = 18  # R


# (auftrag, name, artikelnummer, menge, stck_pal, laenge, breite, hoehe, p_anzahl)
# hoehe == None  → Zelle bleibt leer
DATEN = [
    (110135, "VDL Jansen bv",        "PSCJ003637-0450/001",  2, 350, 1220,  950, 1100, 1),
    (110287, "BITO Lagertechnik",    "PSC2403x1508/001",     2, 100, 2450, 1550, None, 1),
    (109931, "Hellmann Poultry",     "PSC45-36-0416/004",    1, 100, 2420, 1050,  790, 1),
    (110021, "Hellmann Poultry",     "PSC45-36-0416/004",    5, 100, 2420, 1050,  790, 1),
    (109767, "GI-RO Technik GmbH",   "EWM822x880/001",      48,  38,  910,  850, 1300, 2),
    (110017, "Hellmann Poultry",     "PSC45-36-1049/001",    1, 100, 2415,  760,  780, 1),
    (110150, "R. Inauen AG",         "PSV1995x442/001",      1, 100, 2050,  950, None, 1),
    (110203, "Bressers Metaal BV",   "PSV1160x840/001",      2, 125, 1200,  800,  950, 1),
    (110070, "Big Dutchman",         "PSF83-50-4400/001",    9, 112, 1620, 1225, 1340, 1),
    (109971, "VDL Jansen bv",        "PSCJ2.5X50X50ZNA/001", 8, 125, 3050, 1550,  500, 1),
    (110260, "Hellmann Poultry",     "PSC45-36-0868/001",    2, 500, 1180,  800,  930, 1),
    (110224, "Hellmann Poultry",     "PSC45-36-0868/001",    2, 500, 1180,  800,  930, 1),
    (109576, "Big Dutchman",         "PSF83-07-2880/002",   10, 112, 1430, 1200, 1200, 1),
    (110055, "Urban GmbH & Co.",     "PSB1631x441/006",      1,  50, 1680,  500, None, 1),
    (110055, "Urban GmbH & Co.",     "PSB1631x441/004",      2,  50, 1680,  500, None, 1),
]


def baue_real(pfad: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabelle1"

    # Z1 A: Firmenname
    ws.cell(row=1, column=1, value="Draht Müller GmbH & Co. KG")
    # Z2: leer
    # Z3 A: 'verplante Maschinen', K: Datum/Seite
    ws.cell(row=3, column=1, value="verplante Maschinen")
    ws.cell(row=3, column=11, value="Datum '07.05.26' 20:53, Seite 1")
    # Z4: leer
    # Z5: Header (J und K bleiben leer)
    ws.cell(row=5, column=1,  value="Maschine")          # A
    ws.cell(row=5, column=2,  value="KW Fertigung")      # B
    ws.cell(row=5, column=3,  value="Nr.")               # C
    ws.cell(row=5, column=4,  value="KW Lief")           # D
    ws.cell(row=5, column=5,  value="Auftrag")           # E
    ws.cell(row=5, column=6,  value="ANr")               # F
    ws.cell(row=5, column=7,  value="AK")                # G
    ws.cell(row=5, column=8,  value="Name")              # H
    ws.cell(row=5, column=9,  value="Artikelnummer")     # I
    ws.cell(row=5, column=12, value="Masche/Bemerkung")  # L
    ws.cell(row=5, column=13, value="Menge")             # M
    ws.cell(row=5, column=14, value="Stck Pal")          # N
    ws.cell(row=5, column=15, value="P-Länge")           # O
    ws.cell(row=5, column=16, value="P-Breite")          # P
    ws.cell(row=5, column=17, value="P-Höhe")            # Q
    ws.cell(row=5, column=18, value="P-Anzahl")          # R
    # Z6: leer

    # Daten ab Z7, jede zweite Zeile (8,10,12,... leer)
    aktuelle_zeile = 7
    for auftrag, name, artikel, menge, stck, laenge, breite, hoehe, p_anzahl in DATEN:
        ws.cell(row=aktuelle_zeile, column=COL_AUFTRAG,       value=auftrag)
        ws.cell(row=aktuelle_zeile, column=COL_NAME,          value=name)
        ws.cell(row=aktuelle_zeile, column=COL_ARTIKELNUMMER, value=artikel)
        ws.cell(row=aktuelle_zeile, column=COL_MENGE,         value=menge)
        ws.cell(row=aktuelle_zeile, column=COL_STCK_PAL,      value=stck)
        ws.cell(row=aktuelle_zeile, column=COL_LAENGE,        value=laenge)
        ws.cell(row=aktuelle_zeile, column=COL_BREITE,        value=breite)
        if hoehe is not None:
            ws.cell(row=aktuelle_zeile, column=COL_HOEHE, value=hoehe)
        ws.cell(row=aktuelle_zeile, column=COL_P_ANZAHL,      value=p_anzahl)
        aktuelle_zeile += 2  # Leerzeile dazwischen

    pfad.parent.mkdir(parents=True, exist_ok=True)
    wb.save(pfad)


if __name__ == "__main__":
    ziel = Path(__file__).resolve().parent / "draht_mueller_real.xlsx"
    baue_real(ziel)
    print(f"Fixture erstellt: {ziel} ({ziel.stat().st_size} bytes)")
    print(f"15 Datenzeilen, dazwischen Leerzeilen.")
    print(f"Menge-Summe erwartet: {sum(d[3] for d in DATEN)} (Soll: 96)")
    print(f"GI-RO 109767 Menge: {[d[3] for d in DATEN if d[0] == 109767][0]} (Soll: 48)")
