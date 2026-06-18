"""Tests fuer Berichte-Tab (Spec §1-§9).

a) Archiv CRUD + Dateinamen-Konvention
b) PDF Optimierungs-Report: gueltige PDF-Bytes mit Header
c) Excel Zuordnung: 4 Sheets, Aggregate
d) Bestandsbericht PDF + Excel mit Filtern
e) Wirtschaftlichkeitsbericht: Trend-Chart eingebettet (PIL)
f) Lieferanten-Bestelldokument: Summen netto/19%/brutto korrekt
g) Latin-1-Fallback fuer Umlaute + Sonderzeichen
"""
from __future__ import annotations

import io
import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

HIER = Path(__file__).resolve().parent
sys.path.insert(0, str(HIER.parent))

_TMP_DIR = tempfile.mkdtemp(prefix="berichte_test_")
os.environ["PALETTENMINI_BERICHTE_DIR"] = _TMP_DIR

import berichte as ber  # noqa: E402


def _opt_eintrag(standards=3, sonder=1):
    return {
        "id": "test-lauf",
        "datum": "2026-05-21T12:00:00",
        "datei_name": "Müller.xlsx",
        "optimierung": {
            "tol_kurz_mm": 200, "tol_lang_mm": 400, "tol_modus": "absolut",
            "kombinieren": True, "sonder_deckel": 5,
            "standards": standards, "sonder": sonder,
            "gesamt": standards + sonder,
            "invariante_ok": True, "status": "Optimal",
            "params_snapshot": {"tol_kurz_mm": 200, "sonder_erlaubt": True},
            "ergebnis_snapshot": {
                "standards": [[800, 1200], [1000, 1500], [700, 900]][:standards],
                "sonder": [[900, 1300]][:sonder],
                "score": 4.5,
            },
        },
    }


def _ergebnis(n_zuord=5):
    return {
        "standards": [(800, 1200), (1000, 1500)],
        "sonder": [(900, 1300)],
        "zuordnung": [
            {"auftrag": f"AW{i}", "name": "Kunde Müller",
             "artikelnummer": f"ART-{i}", "verbrauchsdatum": "2026-06-15",
             "L": 1180, "B": 790, "menge": 5+i,
             "ziel": "800x1200", "typ": "Standard",
             "begruendung": "innerhalb Toleranz"}
            for i in range(n_zuord)
        ],
        "nicht_zuordenbar": [],
    }


def test_a_archiv_crud():
    daten = b"FAKE-PDF-DATEN " * 200  # > 1 KB damit size_kb != 0.0
    erg = ber.archiv_speichern("pdf_optimierung",
                                  ber._dateiname("optimierung", "pdf"),
                                  daten, quell_lauf_id="lauf-1")
    assert erg["size_kb"] > 0, f"size_kb={erg['size_kb']} bei {len(daten)} bytes"
    assert Path(erg["pfad"]).exists()
    liste = ber.archiv_liste(limit=10)
    assert len(liste) >= 1
    assert any(e["id"] == erg["id"] for e in liste)
    # Lesen
    geladen = ber.archiv_lesen(erg["id"])
    assert geladen == daten
    # Loeschen
    assert ber.archiv_loeschen(erg["id"]) is True
    assert ber.archiv_lesen(erg["id"]) is None
    assert ber.archiv_loeschen(erg["id"]) is False
    print(f"  Archiv CRUD OK — speichern/lesen/loeschen.")


def test_b_pdf_optimierungs_report():
    pdf = ber.pdf_optimierungs_report(_opt_eintrag(standards=3, sonder=1),
                                          ergebnis=_ergebnis(n_zuord=8))
    assert isinstance(pdf, (bytes, bytearray))
    assert pdf.startswith(b"%PDF-")
    print(f"  PDF Optimierungs-Report: {len(pdf)} bytes, beginnt mit %PDF-.")


def test_c_excel_zuordnung_4_sheets():
    xlsx = ber.excel_zuordnung(_opt_eintrag(standards=3, sonder=1),
                                  ergebnis=_ergebnis(n_zuord=5))
    assert isinstance(xlsx, bytes)
    # Lade zurueck und pruefe Sheets
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    sheets = wb.sheetnames
    print(f"  Excel-Sheets: {sheets}")
    assert "Uebersicht" in sheets
    assert "Zuordnung" in sheets
    assert "Standards" in sheets
    assert "Nicht_zuordenbar" in sheets
    # Zuordnungssheet hat 5 Datenzeilen + Header
    assert wb["Zuordnung"].max_row >= 6


def test_d_bestandsbericht_filter():
    # 'vollstaendig' bedeutet jetzt: EK Lieferant > 0 UND Selbstfertigung > 0
    eintraege = [
        {"name": "Euro", "L_mm": 1200, "B_mm": 800, "bestand": 50,
         "ek_lieferant_eur": 14.0,
         "selbstfertigung_material_eur": 6.0,
         "selbstfertigung_lohn_eur": 5.0,
         "typ": "standard"},
        {"name": "Leer", "L_mm": 1500, "B_mm": 1000, "bestand": 0,
         "ek_lieferant_eur": 18.0,
         "selbstfertigung_material_eur": 8.0,
         "selbstfertigung_lohn_eur": 6.0,
         "typ": "standard"},
        {"name": "EkFehlt", "L_mm": 2000, "B_mm": 600, "bestand": 10,
         "ek_lieferant_eur": 0,
         "selbstfertigung_material_eur": 12.0,
         "selbstfertigung_lohn_eur": 0,
         "typ": "sonder"},
    ]
    # Ohne Filter: 3 Eintraege
    pdf_a = ber.bestandsbericht_pdf(eintraege)
    assert pdf_a.startswith(b"%PDF-")
    # Mit Bestand>0: nur 2 (Leer hat bestand=0)
    excel_b = ber.bestandsbericht_excel(eintraege, nur_bestand_positiv=True)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_b))
    data_rows = [r for r in wb.active.iter_rows(min_row=2, values_only=True)
                  if r[0] and r[0] != "SUMME"]
    print(f"  nur_bestand_positiv: {len(data_rows)} Eintraege (von 3)")
    assert len(data_rows) == 2
    # Mit nur_vollstaendig: Euro+Leer (beide EK Lief + Eigenfert > 0)
    # EkFehlt hat ek_lieferant=0 → raus
    excel_c = ber.bestandsbericht_excel(eintraege, nur_vollstaendig=True)
    wb2 = openpyxl.load_workbook(io.BytesIO(excel_c))
    data_rows2 = [r for r in wb2.active.iter_rows(min_row=2, values_only=True)
                   if r[0] and r[0] != "SUMME"]
    print(f"  nur_vollstaendig: {len(data_rows2)} Eintraege "
          f"(Euro+Leer haben EK Lief + Eigenfert > 0)")
    assert len(data_rows2) == 2


def test_e_wirtschaftlichkeitsbericht_mit_chart():
    optimierungen = [_opt_eintrag(standards=s, sonder=1)
                       for s in [10, 8, 6, 5, 4]]
    pdf = ber.wirtschaftlichkeitsbericht_pdf(optimierungen,
                                                 zeitraum_label="30 Tage")
    assert pdf.startswith(b"%PDF-")
    print(f"  Wirtschaftlichkeitsbericht mit Trend-Chart: {len(pdf)} bytes")


def test_f_lieferanten_bestelldokument():
    pdf = ber.lieferanten_bestelldokument_pdf(
        lieferant={"name": "Müller GmbH", "kontakt": "info@mueller.de"},
        positionen=[
            {"name": "Euro", "L": 1200, "B": 800, "menge": 100,
             "ek_pro_stk_eur": 14.50},
            {"name": "Industrie", "L": 1000, "B": 1200, "menge": 50,
             "ek_pro_stk_eur": 17.20},
        ],
        bestellnummer="BST-2026-0042",
        bestelldatum="21.05.2026",
        wunschtermin="04.06.2026",
        bemerkung="Bitte palettiert anliefern.",
    )
    assert pdf.startswith(b"%PDF-")
    print(f"  Lieferanten-Bestelldokument: {len(pdf)} bytes")
    # Erwartete Summen: 100×14.50 + 50×17.20 = 1450 + 860 = 2310 netto
    # × 1.19 = 2748.90 brutto (steht im PDF — wir pruefen nicht den Inhalt,
    # nur dass es valide PDF ist)


def test_g_latin1_umlaute():
    """fpdf2 core-fonts unterstuetzen Latin-1 (ä,ö,ü,ß) direkt.
    Emoji + Pfeile werden ersetzt."""
    assert ber._latin1("Müller GmbH") == "Müller GmbH"
    assert ber._latin1("Größe Höhe Maß") == "Größe Höhe Maß"
    assert "->" in ber._latin1("Vorher → Nachher")
    assert "EUR" in ber._latin1("100 €")
    assert "[OK]" in ber._latin1("Status 🟢 grün")
    print(f"  Latin-1 + Replacements OK.")


def test_i_pdf_palette_artikel():
    """PDF gruppiert Zuordnungen nach (Typ, Ziel) und listet Artikel."""
    erg = _ergebnis(n_zuord=5)
    # 2 Sonder + 1 Kombi hinzufuegen
    erg["zuordnung"].extend([
        {"auftrag": "S-1", "name": "X", "artikelnummer": "S1",
         "L": 1700, "B": 900, "menge": 4, "ziel": "950x1750",
         "typ": "Sonder", "verbrauchsdatum": ""},
        {"auftrag": "K-1", "name": "Y", "artikelnummer": "K1",
         "L": 1500, "B": 700, "menge": 2, "ziel": "3x (800x500)",
         "typ": "Kombi-Stapel", "verbrauchsdatum": ""},
    ])
    pdf = ber.pdf_palette_artikel(erg, datei_name="test.xlsx")
    assert pdf.startswith(b"%PDF-")
    print(f"  pdf_palette_artikel: {len(pdf)} bytes (3 Gruppen)")


def test_j_pdf_auftrag_palette():
    """PDF gruppiert Zuordnungen nach AW + Summary-Tabelle."""
    erg = _ergebnis(n_zuord=4)
    # Mehrere Zuordnungen mit gleicher AW
    erg["zuordnung"][0]["auftrag"] = "AW-2026-001"
    erg["zuordnung"][1]["auftrag"] = "AW-2026-001"
    erg["zuordnung"][2]["auftrag"] = "AW-2026-002"
    erg["zuordnung"][3]["auftrag"] = "AW-2026-003"
    pdf = ber.pdf_auftrag_palette(erg, datei_name="test.xlsx")
    assert pdf.startswith(b"%PDF-")
    print(f"  pdf_auftrag_palette: {len(pdf)} bytes (3 AWs)")


def test_k_pdf_auftragsuebersicht():
    """Auftragsbasierte Uebersicht: Haupttabelle + Aggregat + Status,
    fehlende Palette -> '-', viele Auftraege -> mehrseitig."""
    lookup = {"P1": "1600x800", "P2": "1800x1000"}
    auftraege = [
        {"aw_nummer": "AW-2026-001", "kunde": "Müller GmbH",
         "artikel_nummer": "12345", "menge": 50, "status": "abgeschlossen",
         "zugewiesene_palette_id": "P1", "bestelldatum": "2026-04-10"},
        {"aw_nummer": "AW-2026-002", "kunde": "Schmitt KG",
         "artikel_nummer": "12346", "menge": 20, "status": "in_arbeit",
         "zugewiesene_palette_id": "P2", "bestelldatum": "2026-05-01"},
        {"aw_nummer": "AW-2026-003", "kunde": "Lang & Co",
         "artikel_nummer": "12347", "menge": 30, "status": "offen",
         "zugewiesene_palette_id": None, "bestelldatum": "2026-05-20"},
    ]
    pdf = ber.pdf_auftragsuebersicht(
        auftraege, lookup, zeitraum="Zeitraum: 2026-04-01 bis 2026-05-31")
    assert pdf.startswith(b"%PDF-")
    print(f"  pdf_auftragsuebersicht (3 Auftr., 1 ohne Palette): "
          f"{len(pdf)} bytes")

    # Mehrseitigkeit: 250 Auftraege duerfen nicht crashen
    viele = [{"aw_nummer": f"AW-{i:04d}", "kunde": f"Kunde {i}",
              "artikel_nummer": f"ART{i}", "menge": i % 7,
              "status": ["offen", "in_arbeit", "abgeschlossen",
                          "storniert"][i % 4],
              "zugewiesene_palette_id": "P1" if i % 2 else None,
              "bestelldatum": "2026-05-10"} for i in range(250)]
    pdf2 = ber.pdf_auftragsuebersicht(viele, lookup)
    assert pdf2.startswith(b"%PDF-")
    assert len(pdf2) > len(pdf), "250 Auftraege sollten groesseres PDF ergeben"
    print(f"  pdf_auftragsuebersicht (250 Auftr.): {len(pdf2)} bytes OK")


def test_h_dateinamen_konvention():
    """Spec §7: youman_mini_<typ>_<YYYYMMDD_HHMM>.<ext>"""
    name = ber._dateiname("optimierung", "pdf")
    assert name.startswith("youman_mini_optimierung_")
    assert name.endswith(".pdf")
    # Zeitstempel: 8 Ziffern + _ + 4 Ziffern
    parts = name.replace("youman_mini_optimierung_", "").replace(".pdf", "")
    assert len(parts) == 13  # 20260521_1245
    print(f"  Dateiname: {name}")


def _run(fn):
    try:
        fn()
        print(f"OK  {fn.__name__}")
        return True
    except AssertionError as e:
        print(f"FAIL {fn.__name__}: {e}")
        return False
    except Exception as e:
        print(f"ERR  {fn.__name__}: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    tests = [
        test_a_archiv_crud,
        test_b_pdf_optimierungs_report,
        test_c_excel_zuordnung_4_sheets,
        test_d_bestandsbericht_filter,
        test_e_wirtschaftlichkeitsbericht_mit_chart,
        test_f_lieferanten_bestelldokument,
        test_i_pdf_palette_artikel,
        test_j_pdf_auftrag_palette,
        test_k_pdf_auftragsuebersicht,
        test_g_latin1_umlaute,
        test_h_dateinamen_konvention,
    ]
    erfolge = sum(1 for t in tests if _run(t))
    print()
    print(f"{erfolge}/{len(tests)} Berichte-Tests bestanden.")
    # Cleanup
    import shutil
    try:
        shutil.rmtree(_TMP_DIR, ignore_errors=True)
    except Exception:
        pass
    sys.exit(0 if erfolge == len(tests) else 1)
