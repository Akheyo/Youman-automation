"""Excel-Import/Export und Beispieldaten-Generator.

Liest .xlsx-Dateien mit flexibler Spaltenerkennung (DE/EN-Schreibweisen),
schreibt das Optimierungs-Ergebnis zurück und kann realistische Testdaten
mit Cluster-Verteilung erzeugen.
"""
from __future__ import annotations

import io
import random
from pathlib import Path
from typing import IO, Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from optimizer import OptimierungsErgebnis, Palette


SPALTEN_MAPPING: dict[str, list[str]] = {
    "artikelnummer": [
        "artikelnummer",
        "artikel",
        "artikelnr",
        "artikel-nr",
        "art-nr",
        "artnr",
        "sku",
        "item",
        "item_no",
        "item number",
        "nummer",
    ],
    "laenge": [
        "laenge",
        "länge",
        "lange",
        "length",
        "l",
        "laenge_mm",
        "länge_mm",
        "length_mm",
        "p-laenge",
        "p-länge",
        "p_laenge",
        "p_länge",
        "p.laenge",
        "p.länge",
    ],
    "breite": [
        "breite",
        "width",
        "b",
        "w",
        "breite_mm",
        "width_mm",
        "p-breite",
        "p_breite",
        "p.breite",
    ],
    "hoehe": [
        "hoehe",
        "höhe",
        "height",
        "h",
        "p-hoehe",
        "p-höhe",
        "p_hoehe",
        "p_höhe",
        "p.hoehe",
        "p.höhe",
    ],
    "anzahl": [
        # 'Menge' ist die maßgebliche Spalte für die Palettenanzahl pro
        # Auftrag (Spalte M in der Draht-Müller-Excel). 'P-Anzahl' wird
        # bewusst NICHT mehr gemappt — diese Spalte wird ignoriert.
        "menge",
        "mng",
        "anzahl",
        "anzahl_paletten",
        "anzahl paletten",
        "palettenanzahl",
        "pallet_count",
        "qty_pallets",
        "n_pallets",
        "pallets",
        "qty",
        "quantity",
        "amount",
    ],
    "stueck_pro_palette": [
        "stueckzahl_pro_palette",
        "stückzahl_pro_palette",
        "stueck_pro_palette",
        "stück_pro_palette",
        "stueckzahl pro palette",
        "stk_pro_palette",
        "stck_pal",
        "stck_pal.",
        "stck pal.",
        "stck pal",
        "stk_pal",
        "stk_pal.",
        "items_per_pallet",
        "pieces_per_pallet",
    ],
    "kosten_alt": [
        "palettenkosten",
        "kosten",
        "kosten_alt",
        "palette_cost",
        "cost",
        "preis",
        "price",
    ],
    "kunde": [
        "kunde",
        "name",
        "customer",
        "client",
        "abnehmer",
    ],
    "auftrag": [
        "auftrag",
        "auftragsnummer",
        "auftrag-nr",
        "auftragsnr",
        "order",
        "order_no",
        "ordernumber",
    ],
    "kw_lieferung": [
        "kw_lieferung",
        "kw lief.",
        "kw lief",
        "kw_lief",
        "kw_lief.",
        "lieferwoche",
        "lief-kw",
        "delivery_week",
    ],
    "menge": [
        "menge",
        "stueck",
        "stück",
        "qty",
        "quantity",
        "amount",
    ],
}


def _normalisiere(name: str) -> str:
    return (
        str(name)
        .strip()
        .lower()
        .replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
        .replace("-", "_")
        .replace(" ", "_")
        .replace(".", "_")
    )


def _finde_spalten(header: list[str]) -> dict[str, int]:
    """Ordnet logische Feldnamen den tatsächlichen Spalten-Indizes zu."""
    norm_header = [_normalisiere(h) if h is not None else "" for h in header]
    mapping: dict[str, int] = {}
    for ziel, kandidaten in SPALTEN_MAPPING.items():
        for idx, h in enumerate(norm_header):
            if h in [_normalisiere(k) for k in kandidaten]:
                mapping[ziel] = idx
                break
    return mapping


PFLICHT_SPALTEN = ("artikelnummer", "laenge", "breite")


def _finde_header_zeile(
    rows: list[tuple],
    scan_zeilen: int = 20,
) -> tuple[int, dict[str, int], list[str]]:
    """Sucht die wahrscheinlichste Header-Zeile in den ersten Zeilen.

    Excel-Exporte haben oft Titel-Zeilen ("Firma X", "verplante Maschinen", ...)
    vor der eigentlichen Header-Zeile. Wir scannen die ersten ``scan_zeilen``
    Zeilen, mappen jede gegen unsere SPALTEN_MAPPING und nehmen die Zeile mit
    den meisten Treffern (Pflichtspalten zählen 100x, Optionale je 1x).

    Returns:
        ``(zeilen_index, spalten_mapping, normalisierter_header)``
    """
    best_idx = 0
    best_mapping: dict[str, int] = {}
    best_score = -1
    best_header: list[str] = []
    for idx in range(min(scan_zeilen, len(rows))):
        row = rows[idx]
        header = [str(c) if c is not None else "" for c in row]
        if not any(h.strip() for h in header):
            continue
        mapping = _finde_spalten(header)
        pflicht_score = sum(1 for p in PFLICHT_SPALTEN if p in mapping)
        score = pflicht_score * 100 + len(mapping)
        if score > best_score:
            best_score = score
            best_idx = idx
            best_mapping = mapping
            best_header = header
    return best_idx, best_mapping, best_header


def importiere_excel(
    file_or_buffer: str | Path | IO[bytes],
    sheet_name: str | None = None,
    warnungen: list[str] | None = None,
) -> list[Palette]:
    """Liest eine Palettenliste aus einer Excel-Datei.

    Erkennt die Header-Zeile automatisch (Titel-Zeilen darüber werden
    übersprungen). Pflichtspalten: Artikelnummer, Länge, Breite.

    Palettenanzahl pro Auftrag = Spalte ``Menge`` (Aliase: Mng, Anzahl).
    Wenn Menge leer oder 0 ist, wird der Auftrag mit 0 Paletten erfasst
    (zählt nicht in die Gesamt-Palettenzahl, das Maß zählt aber weiter
    als Variante) und eine Warnung in ``warnungen`` geschrieben.

    Args:
        file_or_buffer: Pfad oder Datei-ähnliches Objekt (z.B. Streamlit-Upload).
        sheet_name: Optionaler Name des Sheets, sonst das erste.
        warnungen: Optionale Liste, in die Diagnose-Meldungen geschrieben
            werden (z.B. Aufträge mit leerer Menge).

    Returns:
        Liste valider ``Palette``-Objekte. Ungültige Zeilen werden
        übersprungen.
    """
    log = warnungen if warnungen is not None else []
    wb = openpyxl.load_workbook(file_or_buffer, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx, spalten, header = _finde_header_zeile(rows)

    fehlend = [p for p in PFLICHT_SPALTEN if p not in spalten]
    if fehlend:
        gezeigt = [h for h in header if h.strip()] or [str(c) for c in rows[0] if c is not None]
        raise ValueError(
            f"Pflichtspalten fehlen in Excel: {', '.join(fehlend)}. "
            f"Erwartet: Artikelnummer, P-Länge, P-Breite (oder Synonyme). "
            f"Gefundene Spalten in Zeile {header_idx + 1}: {gezeigt}"
        )

    def _val(row: tuple, schluessel: str):
        if schluessel not in spalten:
            return None
        idx = spalten[schluessel]
        if idx >= len(row):
            return None
        return row[idx]

    def _str(v) -> str:
        if v is None:
            return ""
        return str(v).strip()

    def _float(v, default: float = 0.0) -> float:
        if v is None or _str(v) == "":
            return default
        try:
            return float(str(v).replace(",", "."))
        except (TypeError, ValueError):
            return default

    def _int(v, default: int = 0) -> int:
        return int(_float(v, default))

    paletten: list[Palette] = []
    for row_nr, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        try:
            artikel = _str(_val(row, "artikelnummer"))
            laenge = _float(_val(row, "laenge"))
            breite = _float(_val(row, "breite"))
            if not artikel or laenge <= 0 or breite <= 0:
                continue

            auftrag_id = _str(_val(row, "auftrag")) or f"Zeile {row_nr}"
            # Palettenanzahl = Wert aus Spalte 'Menge' (Spalte M).
            # Wenn leer oder 0 → Auftrag mit 0 Paletten erfassen + Warnung.
            anzahl = _int(_val(row, "anzahl"))
            stk_pal = _int(_val(row, "stueck_pro_palette"))
            if anzahl <= 0:
                log.append(
                    f"⚠️ Auftrag {auftrag_id}: Spalte 'Menge' leer oder 0 "
                    "— Auftrag zählt 0 Paletten."
                )
                anzahl = 0

            paletten.append(
                Palette(
                    artikelnummer=artikel,
                    laenge=laenge,
                    breite=breite,
                    laenge_original=laenge,
                    breite_original=breite,
                    anzahl=anzahl,
                    kosten_alt=_float(_val(row, "kosten_alt"), 0.0),
                    stueck_pro_palette=stk_pal,
                    hoehe=_float(_val(row, "hoehe")),
                    kunde=_str(_val(row, "kunde")),
                    auftrag=_str(_val(row, "auftrag")),
                    kw_lieferung=_str(_val(row, "kw_lieferung")),
                    menge=anzahl,  # 'Menge' = Palettenanzahl (mit anzahl gleichgesetzt)
                )
            )
        except (TypeError, ValueError):
            continue

    return paletten


def exportiere_ergebnis_excel(ergebnis: OptimierungsErgebnis) -> bytes:
    """Schreibt das Optimierungs-Ergebnis als Excel-Datei (Bytes)."""
    wb = openpyxl.Workbook()
    ws_std = wb.active
    ws_std.title = "Standards"
    header_fill = PatternFill("solid", fgColor="1A2944")
    header_font = Font(color="FFFFFF", bold=True)

    spalten_std = [
        "Standard",
        "Länge (mm)",
        "Breite (mm)",
        "Anzahl Mitglieder",
        "Gesamt Paletten",
    ]
    ws_std.append(spalten_std)
    for col, _ in enumerate(spalten_std, start=1):
        c = ws_std.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for s in ergebnis.standards:
        ws_std.append(
            [
                s.label,
                round(s.laenge),
                round(s.breite),
                len(s.members),
                s.gesamt_anzahl,
            ]
        )

    for col_idx in range(1, len(spalten_std) + 1):
        ws_std.column_dimensions[get_column_letter(col_idx)].width = 22

    ws_det = wb.create_sheet("Zuordnung")
    spalten_det = [
        "Standard",
        "Artikelnummer",
        "Original Länge",
        "Original Breite",
        "Anzahl",
    ]
    ws_det.append(spalten_det)
    for col, _ in enumerate(spalten_det, start=1):
        c = ws_det.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for s in ergebnis.standards:
        for m in s.members:
            ws_det.append(
                [
                    s.label, m.artikelnummer,
                    m.laenge_original or m.laenge,
                    m.breite_original or m.breite,
                    m.anzahl,
                ]
            )
    for col_idx in range(1, len(spalten_det) + 1):
        ws_det.column_dimensions[get_column_letter(col_idx)].width = 22

    if ergebnis.kombinationen:
        ws_kombi = wb.create_sheet("Kombinationen")
        spalten_kombi = [
            "Artikelnummer",
            "Original Länge",
            "Original Breite",
            "Zusammensetzung",
            "Richtung",
            "Anzahl Standards",
        ]
        ws_kombi.append(spalten_kombi)
        for col, _ in enumerate(spalten_kombi, start=1):
            c = ws_kombi.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
        for k in ergebnis.kombinationen:
            zusammensetzung = " + ".join(s.label for s in k.standards)
            richtung = "längs" if k.richtung == "laenge" else "quer"
            ws_kombi.append(
                [
                    k.palette.artikelnummer,
                    k.palette.laenge_original or k.palette.laenge,
                    k.palette.breite_original or k.palette.breite,
                    zusammensetzung,
                    richtung,
                    len(k.standards),
                ]
            )
        for col_idx in range(1, len(spalten_kombi) + 1):
            ws_kombi.column_dimensions[get_column_letter(col_idx)].width = 26

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def erstelle_beispiel_excel(pfad: str | Path, n_artikel: int = 80) -> Path:
    """Erzeugt eine realistische Beispiel-Palettenliste.

    Die Größen werden aus einigen Cluster-Zentren mit ±20mm Rauschen gezogen,
    sodass die Optimierung sinnvolle Standards finden kann.
    """
    rnd = random.Random(42)
    cluster: list[tuple[float, float]] = [
        (1500, 700),
        (1800, 1000),
        (1200, 800),
        (1600, 800),
        (2000, 1200),
        (1000, 600),
    ]

    rows: list[tuple[str, int, int, int, int, float]] = []
    for i in range(n_artikel):
        c = rnd.choice(cluster)
        l = c[0] + rnd.randint(-20, 20)
        b = c[1] + rnd.randint(-15, 15)
        anzahl = rnd.randint(5, 50)
        stueck = rnd.choice([20, 25, 30, 40, 50])
        kosten_alt = round(rnd.uniform(12.0, 16.0), 2)
        artnr = f"ART-{1000 + i:04d}"
        rows.append((artnr, int(l), int(b), anzahl, stueck, kosten_alt))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Palettenliste"

    header = [
        "Artikelnummer",
        "Laenge",
        "Breite",
        "Benoetigte_Paletten",
        "Stueckzahl_pro_Palette",
        "Palettenkosten",
    ]
    ws.append(header)

    header_fill = PatternFill("solid", fgColor="1A2944")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(header, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    for r in rows:
        ws.append(list(r))

    pfad = Path(pfad)
    pfad.parent.mkdir(parents=True, exist_ok=True)
    wb.save(pfad)
    return pfad


def paletten_aus_iterable(
    eintraege: Iterable[dict],
    default_kosten: float = 14.0,
) -> list[Palette]:
    """Hilfsfunktion: erzeugt Paletten aus einer Liste von Dicts."""
    out: list[Palette] = []
    for e in eintraege:
        try:
            out.append(
                Palette(
                    artikelnummer=str(e["artikelnummer"]),
                    laenge=float(e["laenge"]),
                    breite=float(e["breite"]),
                    anzahl=int(e.get("anzahl", 1)),
                    kosten_alt=float(e.get("kosten_alt", default_kosten)),
                    stueck_pro_palette=int(e.get("stueck_pro_palette", 0)),
                )
            )
        except (KeyError, TypeError, ValueError):
            continue
    return out


def exportiere_zuordnung_excel(rows: list[dict]) -> bytes:
    """Exportiert die Auftrag→Standard-Zuordnungstabelle als Excel-Datei.

    Spalten entsprechen der Detail-Tabelle in der UI (eine Zeile pro
    Auftragsposition mit Status-Spalte und Abweichung in mm).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zuordnung"

    spalten = [
        "Auftrag", "Kunde", "Artikelnummer",
        "Original L", "Original B", "Original H",
        "Menge (Paletten)", "Stk/Pal",
        "Zugewiesener Standard", "Abw. L (mm)", "Abw. B (mm)",
        "Abw. H (mm)", "Größte Abw. (mm)", "Status",
    ]
    ws.append(spalten)

    header_fill = PatternFill("solid", fgColor="1A2944")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(spalten, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    status_label = {
        "standard": "Standard",
        "kombination": "Kombination",
        "sonder": "Sonderpalette",
    }
    for r in rows:
        ws.append([
            r.get("auftrag", ""),
            r.get("kunde", ""),
            r.get("artikelnummer", ""),
            int(r["original_l"]) if r.get("original_l") else "",
            int(r["original_b"]) if r.get("original_b") else "",
            int(r["original_h"]) if r.get("original_h") else "",
            r.get("anzahl", 0),
            r.get("stk_pal", 0),
            r.get("standard_label", ""),
            r.get("abw_l", 0),
            r.get("abw_b", 0),
            r.get("abw_h", 0),
            r.get("abw_max", 0),
            status_label.get(r.get("status", ""), r.get("status", "")),
        ])

    for col_idx in range(1, len(spalten) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
